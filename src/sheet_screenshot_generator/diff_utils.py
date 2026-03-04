from pathlib import Path
from typing import List, Dict, Any, Sequence, Optional, Set

import logging
import openpyxl
import json
from openpyxl.chartsheet import Chartsheet

from src.label_file_diff import build_workbook_diff

# Cache for openpyxl Workbooks (data_only=False)
_WB_NORMAL_CACHE: Dict[str, openpyxl.Workbook] = {}


def _get_wb_normal(path: Path) -> openpyxl.Workbook:
    """
    Cached openpyxl.load_workbook(data_only=False).

    The same path is only loaded once and reused afterwards.
    """
    key = str(path)
    wb = _WB_NORMAL_CACHE.get(key)
    if wb is not None:
        return wb

    wb = openpyxl.load_workbook(str(path), data_only=False)
    _WB_NORMAL_CACHE[key] = wb
    return wb


def _compute_focus_sheets_for_modify(
    input_xlsx: Path, answer_xlsx: Path
) -> Optional[Set[str]]:
    """
    For "modify" tasks:

    - If the sheet set of answer.xlsx is a subset of input.xlsx's sheets,
      return that subset (focus only on those sheets);
    - Otherwise return None (meaning use all sheets).
    """
    try:
        wb_input = _get_wb_normal(input_xlsx)
        wb_answer = _get_wb_normal(answer_xlsx)
    except Exception as e:
        print(f"   !! Failed to open workbooks for subset-sheet logic: {e}")
        return None

    input_set = set(wb_input.sheetnames)
    answer_set = set(wb_answer.sheetnames)

    if answer_set and answer_set.issubset(input_set):
        print(
            "   >> answer sheets are subset of input sheets, "
            f"only focus on: {sorted(answer_set)}"
        )
        return answer_set

    print(
        "   >> answer sheets are NOT subset of input sheets; using all sheets. "
        f"input={sorted(input_set)}, answer={sorted(answer_set)}"
    )
    return None


def _sheet_chart_signature(ws) -> List[Dict[str, Any]]:
    """
    Build a "chart signature" for one sheet.

    We keep only comparison-relevant fields:
    type, title, position, series ranges, etc.
    """
    sig_list: List[Dict[str, Any]] = []

    for ch in getattr(ws, "_charts", []):
        ch_type = type(ch).__name__

        title = ""
        try:
            if getattr(ch, "title", None):
                title = str(ch.title)
        except Exception:
            pass

        position = ""
        try:
            if getattr(ch, "anchor", None):
                position = str(ch.anchor)
        except Exception:
            pass

        series_info: List[Dict[str, str]] = []
        try:
            if getattr(ch, "series", None):
                for s in ch.series:
                    try:
                        v = getattr(s, "values", None)
                        xv = getattr(s, "xvalues", None)
                        series_info.append(
                            {
                                "values": "" if v is None else str(v),
                                "xvalues": "" if xv is None else str(xv),
                            }
                        )
                    except Exception:
                        continue
        except Exception:
            pass

        sig_list.append(
            {
                "type": ch_type,
                "title": title,
                "position": position,
                "series": series_info,
            }
        )

    sig_list.sort(
        key=lambda c: (c["type"], c["title"], c["position"], str(c["series"]))
    )
    return sig_list


def get_chart_diff_sheet_names(
    wb1: Path,
    wb2: Path,
    focus_sheets: Optional[Sequence[str]] = None,
) -> List[str]:
    """
    Compare only charts and return sheets with chart differences.

    Rules:
    - Only compare sheets that exist in both workbooks (same sheet name);
    - If chart count differs or any chart signature differs, the sheet is considered changed;
    - Even if cell values are identical, chart-only changes still mark the sheet as changed.
    """
    changed: List[str] = []

    try:
        wb1_obj = _get_wb_normal(wb1)
        wb2_obj = _get_wb_normal(wb2)
    except Exception as e:
        print(f"   !! Failed to open workbooks for chart diff: {e}")
        return changed

    names1 = set(wb1_obj.sheetnames)
    names2 = set(wb2_obj.sheetnames)
    common = names1 & names2
    if focus_sheets:
        common &= set(focus_sheets)

    for name in common:
        ws1 = wb1_obj[name]
        ws2 = wb2_obj[name]

        sig1 = _sheet_chart_signature(ws1)
        sig2 = _sheet_chart_signature(ws2)

        if sig1 != sig2:
            changed.append(name)

    return sorted(set(changed))


def _cell_style_key(cell) -> tuple:
    """
    Compress a cell's visible style into a tuple for comparison.

    We only keep fields that affect visible appearance (border, fill,
    font, alignment, number format), to avoid openpyxl object noise.
    """
    b = cell.border
    f = cell.fill
    ft = cell.font
    al = cell.alignment

    border_key = (
        b.left.style,
        getattr(b.left.color, "rgb", None),
        b.right.style,
        getattr(b.right.color, "rgb", None),
        b.top.style,
        getattr(b.top.color, "rgb", None),
        b.bottom.style,
        getattr(b.bottom.color, "rgb", None),
    )

    fill_key = (
        f.patternType,
        getattr(f.fgColor, "type", None),
        getattr(f.fgColor, "rgb", None),
        getattr(f.bgColor, "type", None),
        getattr(f.bgColor, "rgb", None),
    )

    font_key = (
        ft.name,
        ft.sz,
        ft.bold,
        ft.italic,
        ft.underline,
        getattr(ft.color, "rgb", None),
    )

    align_key = (
        al.horizontal,
        al.vertical,
        al.wrapText,
        al.shrinkToFit,
        al.indent,
        al.textRotation,
    )

    numfmt_key = cell.number_format

    return border_key, fill_key, font_key, align_key, numfmt_key


def get_formatting_diff_sheet_names(
    wb1: Path,
    wb2: Path,
    max_rows: int = 500,
    max_cols: int = 100,
    focus_sheets: Optional[Sequence[str]] = None,
) -> List[str]:
    """
    Compare only cell formatting (border, fill, font, alignment, number_format)
    and return sheet names with formatting differences.

    For performance, only compare in [1 .. max_rows] × [1 .. max_cols].
    Chartsheets are skipped with a warning.
    """
    changed: List[str] = []

    try:
        wb1_obj = _get_wb_normal(wb1)
        wb2_obj = _get_wb_normal(wb2)
    except Exception as e:
        print(f"   !! Failed to open workbooks for formatting diff: {e}")
        return changed

    names1 = set(wb1_obj.sheetnames)
    names2 = set(wb2_obj.sheetnames)
    common = names1 & names2
    if focus_sheets:
        common &= set(focus_sheets)

    for name in common:
        ws1 = wb1_obj[name]
        ws2 = wb2_obj[name]

        if isinstance(ws1, Chartsheet) or isinstance(ws2, Chartsheet):
            print(
                f"   !! Skip formatting diff for chart sheet '{name}' "
                f"in {wb1.name} vs {wb2.name}: "
                f"Chartsheet has no max_row/max_column (no cell grid)."
            )
            continue

        if not all(
            hasattr(ws, "max_row") and hasattr(ws, "max_column")
            for ws in (ws1, ws2)
        ):
            print(
                f"   !! Skip formatting diff for non-worksheet sheet '{name}' "
                f"in {wb1.name} vs {wb2.name}: "
                f"sheet object has no max_row/max_column."
            )
            continue

        rows = min(max(ws1.max_row, ws2.max_row), max_rows)
        cols = min(max(ws1.max_column, ws2.max_column), max_cols)

        has_diff = False
        for r in range(1, rows + 1):
            if has_diff:
                break
            for c in range(1, cols + 1):
                c1 = ws1.cell(row=r, column=c)
                c2 = ws2.cell(row=r, column=c)

                if _cell_style_key(c1) != _cell_style_key(c2):
                    has_diff = True
                    break

        if has_diff:
            changed.append(name)

    return sorted(set(changed))


def get_changed_sheet_names_for_pair(
    wb1: Path,
    wb2: Path,
    include_formatting: bool = True,
    focus_sheets: Optional[Sequence[str]] = None,
) -> List[str]:
    """
    Use build_workbook_diff + chart (+ optional formatting) to determine
    which sheets are changed between two workbooks.

    A sheet is considered changed if:
    - There is any cell text/formula difference (diff non-empty),
    - OR the chart signature differs,
    - OR (when include_formatting=True) there is a formatting difference,
    - OR the sheet exists only on one side (structural change).
    """
    # 1) cell / formula differences
    try:
        diff_full = build_workbook_diff(wb1, wb2)
        sheets_diff = diff_full.get("sheets_diff", {}) or {}
        cell_changed = {
            sheet_name
            for sheet_name, sheet_info in sheets_diff.items()
            if sheet_info.get("diff")
        }
    except Exception as e:
        print(f"   !! build_workbook_diff failed for {wb1.name} vs {wb2.name}: {e}")
        diff_full = {}
        cell_changed = set()

    if focus_sheets:
        focus_set = set(focus_sheets)
        cell_changed &= focus_set

    # 2) chart differences
    chart_changed = set(
        get_chart_diff_sheet_names(wb1, wb2, focus_sheets=focus_sheets)
    )

    # 3) formatting differences (optional)
    if include_formatting:
        formatting_changed = set(
            get_formatting_diff_sheet_names(wb1, wb2, focus_sheets=focus_sheets)
        )
    else:
        formatting_changed = set()

    # 4) structural differences: sheets only on one side
    before_sheet_names = set(diff_full.get("before_sheets", {}).keys())
    after_sheet_names = set(diff_full.get("after_sheets", {}).keys())
    structural_changed = (before_sheet_names - after_sheet_names) | (
        after_sheet_names - before_sheet_names
    )
    if focus_sheets:
        focus_set = set(focus_sheets)
        structural_changed &= focus_set

    print(f"   >> cell_changed: {sorted(cell_changed)}")
    print(f"   >> chart_changed: {sorted(chart_changed)}")
    if include_formatting:
        print(f"   >> formatting_changed: {sorted(formatting_changed)}")
    print(f"   >> structural_changed: {sorted(structural_changed)}")

    all_changed = sorted(
        cell_changed | chart_changed | formatting_changed | structural_changed
    )
    return all_changed


def _get_file_mode_from_property(case_path: Path) -> str:
    prop_path = case_path / "property.json"
    default_mode = "modify"

    if not prop_path.is_file():
        return default_mode

    try:
        with prop_path.open("r", encoding="utf-8") as f:
            data = json.load(f)
    except Exception as e:
        print(f"   !! Failed to load property.json in {case_path.name}: {e}")
        return default_mode

    mode = data.get("file_mode") or data.get("task_classification") or default_mode
    mode = str(mode).lower()

    if mode not in ("modify", "generate", "qa"):
        print(f"   !! Unknown file_mode '{mode}' in {prop_path}, fallback to 'modify'")
        return default_mode

    return mode


def _compute_focus_sheets_for_modify(input_xlsx: Path, answer_xlsx: Path) -> Optional[Set[str]]:
    try:
        wb_input = _get_wb_normal(input_xlsx)
        wb_answer = _get_wb_normal(answer_xlsx)
    except Exception as e:
        print(f"   !! Failed to open workbooks for subset-sheet logic: {e}")
        return None

    input_set = set(wb_input.sheetnames)
    answer_set = set(wb_answer.sheetnames)

    if answer_set and answer_set.issubset(input_set):
        print(
            f"   >> answer sheets are subset of input sheets, "
            f"only focus on: {sorted(answer_set)}"
        )
        return answer_set

    print(
        f"   >> answer sheets are NOT subset of input sheets; using all sheets. "
        f"input={sorted(input_set)}, answer={sorted(answer_set)}"
    )
    return None


def get_changed_sheet_names_for_pair(
    wb1: Path,
    wb2: Path,
    include_formatting: bool = True,
    focus_sheets: Optional[Sequence[str]] = None,
) -> List[str]:

    # 1) / 
    try:
        diff_full = build_workbook_diff(wb1, wb2)
        sheets_diff = diff_full.get("sheets_diff", {}) or {}
        cell_changed = {
            sheet_name
            for sheet_name, sheet_info in sheets_diff.items()
            if sheet_info.get("diff")
        }
    except Exception as e:
        print(f"   !! build_workbook_diff failed for {wb1.name} vs {wb2.name}: {e}")
        diff_full = {}
        cell_changed = set()

    # focus_sheets， cell diff 
    if focus_sheets:
        focus_set = set(focus_sheets)
        cell_changed &= focus_set

    # 2) （） 
    chart_changed = set(get_chart_diff_sheet_names(wb1, wb2, focus_sheets=focus_sheets))

    # 3) ：（） 
    if include_formatting:
        formatting_changed = set(
            get_formatting_diff_sheet_names(wb1, wb2, focus_sheets=focus_sheets)
        )
    else:
        formatting_changed = set()

    # 4) ： sheet， 
    before_sheet_names = set(diff_full.get("before_sheets", {}).keys())
    after_sheet_names = set(diff_full.get("after_sheets", {}).keys())
    structural_changed = (before_sheet_names - after_sheet_names) | (after_sheet_names - before_sheet_names)
    if focus_sheets:
        focus_set = set(focus_sheets)
        structural_changed &= focus_set

    print(f"   >> cell_changed: {sorted(cell_changed)}")
    print(f"   >> chart_changed: {sorted(chart_changed)}")
    if include_formatting:
        print(f"   >> formatting_changed: {sorted(formatting_changed)}")
    print(f"   >> structural_changed: {sorted(structural_changed)}")

    all_changed = sorted(
        cell_changed | chart_changed | formatting_changed | structural_changed
    )
    return all_changed

def get_formatting_diff_sheet_names(
    wb1: Path,
    wb2: Path,
    max_rows: int = 500,
    max_cols: int = 100,
    focus_sheets: Optional[Sequence[str]] = None,
) -> List[str]:

    changed: List[str] = []

    try:
        wb1_obj = _get_wb_normal(wb1)
        wb2_obj = _get_wb_normal(wb2)
    except Exception as e:
        print(f"   !! Failed to open workbooks for formatting diff: {e}")
        return changed


    names1 = set(wb1_obj.sheetnames)
    names2 = set(wb2_obj.sheetnames)
    common = names1 & names2
    if focus_sheets:
        common &= set(focus_sheets)

    for name in common:
        ws1 = wb1_obj[name]
        ws2 = wb2_obj[name]

        # === ：ChartSheet / ， === 
        if isinstance(ws1, Chartsheet) or isinstance(ws2, Chartsheet):
            print(
                f"   !! Skip formatting diff for chart sheet '{name}' "
                f"in {wb1.name} vs {wb2.name}: "
                f"Chartsheet has no max_row/max_column (no cell grid)."
            )
            continue

        
        if not hasattr(ws1, "max_row") or not hasattr(ws1, "max_column") \
           or not hasattr(ws2, "max_row") or not hasattr(ws2, "max_column"):
            print(
                f"   !! Skip formatting diff for non-worksheet sheet '{name}' "
                f"in {wb1.name} vs {wb2.name}: "
                f"sheet object has no max_row/max_column."
            )
            continue
        # === ===

        # max_rows × max_cols 
        rows = min(max(ws1.max_row, ws2.max_row), max_rows)
        cols = min(max(ws1.max_column, ws2.max_column), max_cols)

        has_diff = False

        for r in range(1, rows + 1):
            if has_diff:
                break
            for c in range(1, cols + 1):
                c1 = ws1.cell(row=r, column=c)
                c2 = ws2.cell(row=r, column=c)

                # “” key， 
                if _cell_style_key(c1) != _cell_style_key(c2):
                    has_diff = True
                    break

        if has_diff:
            changed.append(name)

    return sorted(set(changed))