# gpt_judge_context.py
"""
Build per-case judge contexts (modify / generate / QA).
"""

import json
from pathlib import Path
from typing import Dict, Any, List, Optional, Tuple, Set, Sequence

import pandas as pd
import re

from src.config import PROMPT_MAX_CHARS, USE_GPT_HEADER_REGION, USE_GPT_SHEET_SELECTION
from src.label_file_diff import build_workbook_diff
from .gpt_judge_prompts import (
    JUDGE_PROMPT_MODIFY,
    JUDGE_PROMPT_GENERATE,
    JUDGE_PROMPT_QA,
    build_prompt_for_case
)
from .gpt_judge_excel import (
    build_snapshot_full,
    summarize_diff_only,
    compress_diff,
    collect_changed_rows,
    collect_changed_cols,
    build_before_snapshot_rich,
    build_before_snapshot_simple,
    build_compressed_before_rich,
    build_compressed_before_simple,
    build_before_snapshot_gpt_cells,
    build_compressed_before_gpt_cells,
)
from .gpt_judge_helpers import detect_important_cells_with_gpt, call_gpt_select_sheets
from .gpt_judge_excel import parse_addr  # if needed
from openpyxl import load_workbook


MAX_PROMPT_CHARS = PROMPT_MAX_CHARS


def truncate_prompt_to_max(prompt: str, max_chars: int) -> str:
    """Truncate long prompts and append an explicit truncation marker."""
    if max_chars <= 0:
        return ""
    if len(prompt) <= max_chars:
        return prompt
    suffix = "\n\n...[PROMPT TRUNCATED DUE TO LENGTH LIMIT]"
    if len(suffix) >= max_chars:
        return prompt[:max_chars]
    keep = max_chars - len(suffix)
    return prompt[:keep] + suffix


def read_optional_text_file(path: Path, label: str = "", max_chars: int = 300000) -> str:
    if not path.is_file():
        return ""
    try:
        text = path.read_text(encoding="utf-8", errors="ignore")
    except Exception:
        return ""

    if len(text) > max_chars:
        omitted = len(text) - max_chars
        suffix = label or path.name
        text = text[:max_chars] + f"\n...[{suffix} truncated, {omitted} chars omitted]"
    return text



def load_text_image_sequence_from_folder(folder: Path) -> Tuple[str, List[Path]]:
    if not folder.is_dir():
        return "", []

    text_files: Dict[int, Path] = {}
    img_files: Dict[int, Path] = {}

    for p in folder.iterdir():
        if not p.is_file():
            continue
        name = p.name
        lower = name.lower()

        m_text = re.match(r"text(\d+)\.txt$", lower)
        if m_text:
            idx = int(m_text.group(1))
            text_files[idx] = p
            continue

        m_pic = re.match(r"pic(\d+)\.(png|jpg|jpeg|bmp)$", lower)
        if m_pic:
            idx = int(m_pic.group(1))
            img_files[idx] = p
            continue

    if not text_files and not img_files:
        return "", []

    max_idx = max(list(text_files.keys() or [0]) + list(img_files.keys() or [0]))

    # 1..max_idx ， textN picN 
    segments: List[Tuple[str, Path]] = []
    for i in range(1, max_idx + 1):
        if i in text_files:
            segments.append(("text", text_files[i]))
        if i in img_files:
            segments.append(("image", img_files[i]))

    pieces: List[str] = []
    images: List[Path] = []
    seg_no = 0

    for kind, p in segments:
        seg_no += 1
        if kind == "text":
            try:
                content = p.read_text(encoding="utf-8")
            except UnicodeDecodeError:
                content = p.read_text(encoding="utf-8", errors="ignore")
            pieces.append(f"[TEXT segment {seg_no}]")
            pieces.append(content)
        else:
            images.append(p)
            pieces.append(f"[IMAGE segment {seg_no}: {p.name}]")

    combined = "\n\n".join(pieces)
    return combined, images


def read_task_classification(case_dir: Path) -> str:
    prop_path = case_dir / "property.json"
    if not prop_path.is_file():
        return "modify"

    try:
        data = json.loads(prop_path.read_text(encoding="utf-8", errors="ignore"))
        tc = str(data.get("task_classification", "modify")).strip().lower()
        if tc not in ("modify", "generate", "qa"):
            return "modify"
        return tc
    except Exception:
        return "modify"

def build_case_judge_context(
    input_path: Path,
    answer_path: Path,
    output_path: Path,
    query_text: str,
    file_mode: str = "modify",
) -> Dict[str, Any]:
    file_mode = (file_mode or "modify").lower()

    # ========= ：（“ answer sheet ”） ========= 
    if file_mode == "modify":
        case_dir = input_path.parent  # / property.json 

        # input.txt / answer.txt / output.txt（） 
        input_txt = read_optional_text_file(case_dir / "input.txt", label="input.txt")
        answer_txt = read_optional_text_file(case_dir / "answer.txt", label="answer.txt")
        output_txt = read_optional_text_file(case_dir / "output.txt", label="output.txt")
        
        
        # ---------- answer sheet input ---------- 
        focus_sheets: Optional[Set[str]] = None
        try:
            wb_input = load_workbook(filename=str(input_path), data_only=False)
            wb_answer = load_workbook(filename=str(answer_path), data_only=False)

            input_sheet_names = [ws.title for ws in wb_input.worksheets]
            answer_sheet_names = [ws.title for ws in wb_answer.worksheets]

            input_set = set(input_sheet_names)
            answer_set = set(answer_sheet_names)

            # answer sheet input sheet ，“ sheet” 
            if answer_set and answer_set.issubset(input_set):
                focus_sheets = answer_set
                print(
                    f"[MODIFY] answer sheets are subset of input sheets, "
                    f"only focusing on sheets: {sorted(focus_sheets)}"
                )
            else:
                print(
                    f"[MODIFY] answer sheets are NOT a subset of input sheets, "
                    f"will use all sheets normally. "
                    f"input={sorted(input_set)}, answer={sorted(answer_set)}"
                )
        except Exception as e:
            # sheet ， 
            print(f"[MODIFY][WARN] failed to inspect workbook sheets for subset logic: {e}")
            focus_sheets = None

        # ---------- input→answer diff ---------- 
        diff_answer_full = build_workbook_diff(input_path, answer_path)
        diff_output_full = build_workbook_diff(input_path, output_path)

        # focus_sheets， diff sheet 
        if focus_sheets:
            # 1) diff_answer_full sheets_diff / before_sheets 
            sa = diff_answer_full.get("sheets_diff", {}) or {}
            ba = diff_answer_full.get("before_sheets", {}) or {}

            sa_filtered = {k: v for k, v in sa.items() if k in focus_sheets}
            ba_filtered = {k: v for k, v in ba.items() if k in focus_sheets}

            diff_answer_full = dict(diff_answer_full)
            diff_answer_full["sheets_diff"] = sa_filtered
            diff_answer_full["before_sheets"] = ba_filtered

            # 2) diff_output_full： 
            # - focus_sheets（ answer ） sheet； 
            # - “ output sheet” 
            # （before 、after ）， sheet 。 
            so = diff_output_full.get("sheets_diff", {}) or {}
            bo = diff_output_full.get("before_sheets", {}) or {}

            so_filtered: Dict[str, Any] = {}
            bo_filtered: Dict[str, Any] = {}

            for sheet_name, info in so.items():
                # before/after， sheets_diff before/after 
                before_items = bo.get(sheet_name)
                if before_items is None:
                    # before_sheets ， sheets_diff "before" 
                    before_items = info.get("before", []) or []
                after_items = info.get("after", []) or []

                if sheet_name in focus_sheets:
                    # answer sheet， 
                    so_filtered[sheet_name] = info
                    bo_filtered[sheet_name] = before_items
                    continue

                # answer sheet： 
                # before 、after ⇒ output sheet， 
                # （ input / answer ），。 
                if not before_items and after_items:
                    so_filtered[sheet_name] = info
                    bo_filtered[sheet_name] = before_items

            diff_output_full = dict(diff_output_full)
            diff_output_full["sheets_diff"] = so_filtered
            if bo:
                diff_output_full["before_sheets"] = bo_filtered


            print(
                f"[MODIFY] filtered diff to focus sheets: {sorted(focus_sheets)}; "
                f"answer_sheets_diff={list(sa_filtered.keys())}, "
                f"output_sheets_diff={list(so_filtered.keys())}"
            )

        # BEFORE “ diff_answer_full.before_sheets” 
        before_sheets = diff_answer_full.get("before_sheets", {}) or {}
        # （ focus_sheets None， diff sheet，，） 

        # diff （ diff_*） 
        gt_summary_full = summarize_diff_only(diff_answer_full, max_chars=500000)
        out_summary_full = summarize_diff_only(diff_output_full, max_chars=500000)

        
        input_info_parts: List[str] = [
            "There is one main original workbook: input.xlsx. "
            "The BEFORE snapshots are taken from input.xlsx for selected rows and columns."
        ]

        # txt， 
        if input_txt:
            input_info_parts.append("## In addition to other content, the input also includes the following text:\n" + input_txt)
        if answer_txt:
            input_info_parts.append("## In addition to other content, the answer also includes the following text:\n" + answer_txt)
        if output_txt:
            input_info_parts.append("## In addition to other content, the output also includes the following text: \n" + output_txt)

        input_info = "\n\n".join(input_info_parts)


        # /（ diff ） 
        changed_rows_answer = collect_changed_rows(diff_answer_full)
        changed_rows_output = collect_changed_rows(diff_output_full)
        changed_cols_answer = collect_changed_cols(diff_answer_full)
        changed_cols_output = collect_changed_cols(diff_output_full)

        # ============= GPT ============= 
        if USE_GPT_HEADER_REGION:
            try:
                # input before_sheets sheet 
                all_sheet_names = sorted(before_sheets.keys())

                if not all_sheet_names:
                    raise ValueError("no sheets found in before_sheets (after possible subset filtering)")

                print("[GPT-HEADER] USE_GPT_HEADER_REGION=True, candidate sheets =", all_sheet_names)

                # ---------- ：“ sheet” ---------- 
                target_sheets: Sequence[str] = all_sheet_names
                sheet_select_debug: Optional[Dict[str, Any]] = None

                if USE_GPT_SHEET_SELECTION:
                    print("[GPT-SHEET] USE_GPT_SHEET_SELECTION=True, asking GPT to pick important sheets...")
                    selected_sheets, sheet_select_debug = call_gpt_select_sheets(
                        query=query_text,
                        sheet_names=all_sheet_names,
                        case_name=case_dir.name,
                    )

                    if selected_sheets:
                        target_sheets = selected_sheets
                        print(f"[GPT-SHEET] GPT selected sheets: {target_sheets}")
                    else:
                        # / ： sheet 
                        target_sheets = all_sheet_names
                        print("[GPT-SHEET] no valid sheets returned, fallback to ALL sheets")
                else:
                    print("[GPT-SHEET] USE_GPT_SHEET_SELECTION=False, use ALL sheets")

                # ---------- ： target_sheets HEADER_REGION ---------- 
                important_cells, header_debug = detect_important_cells_with_gpt(
                    case_dir=case_dir,
                    query_text=query_text,
                    before_sheets=before_sheets,
                    target_sheets=target_sheets,
                    input_workbook=input_path,
                )

                # sheet header debug 
                gpt_debug = {
                    "sheet_selection": sheet_select_debug,
                    "header_regions": header_debug,
                }

                if important_cells:
                    # GPT BEFORE JSON 
                    before_snapshot_gpt = build_before_snapshot_gpt_cells(
                        before_sheets=before_sheets,
                        important_cells=important_cells,
                        max_chars=500000,
                    )
                    before_compact_gpt = build_compressed_before_gpt_cells(
                        before_sheets=before_sheets,
                        important_cells=important_cells,
                    )

                    prompt_gpt = build_prompt_for_case(
                        input_info=input_info,
                        before_snapshot=before_snapshot_gpt,
                        query=query_text,
                        gt_summary=gt_summary_full,
                        generated_text=out_summary_full,
                    )
                    prompt_gpt_chars = len(prompt_gpt)

                    if prompt_gpt_chars <= PROMPT_MAX_CHARS:
                        bundle_gpt = {
                            "input_file": str(input_path),
                            "answer_file": str(answer_path),
                            "output_file": str(output_path),
                            "before": before_compact_gpt,
                            "diff_answer": compress_diff(diff_answer_full),
                            "diff_output": compress_diff(diff_output_full),
                            "mode": "gpt_cells",
                            "task_classification": "modify",
                            # GPT bundle ， 
                            "gpt_header_regions_raw": gpt_debug,
                        }

                        print("=== COMPACT CASE BUNDLE (GPT cells, modify, sheet_selection) ===")

                        return {
                            "bundle": bundle_gpt,
                            "before_snapshot": before_snapshot_gpt,
                            "gt_summary": gt_summary_full,
                            "out_summary": out_summary_full,
                            "prompt": prompt_gpt,
                            "prompt_chars": prompt_gpt_chars,
                            "task_classification": "modify",
                            "gpt_header_regions_raw": gpt_debug,
                        }
                    else:
                        print(
                            f"[GPT-HEADER] prompt too long ({prompt_gpt_chars} chars) "
                            f"when using GPT cells, fallback to original logic."
                        )
                else:
                    print(
                        "[GPT-HEADER] important_cells is empty, "
                        "fallback to original rich/simple logic."
                    )

            except Exception as e:
                print(f"[GPT-HEADER][WARN] exception in GPT header region flow: {e}")
                # fallback 

        # ============= FALLBACK： rich/simple （） ============= 

        # ---------- RICH ---------- 
        before_snapshot_rich = build_before_snapshot_rich(
            before_sheets,
            changed_rows_answer,
            changed_rows_output,
            changed_cols_answer,
            changed_cols_output,
            max_rows_front=10,
            max_rows_back=10,
            max_chars=500000,
        )
        before_compact_rich = build_compressed_before_rich(
            before_sheets,
            changed_rows_answer,
            changed_rows_output,
            changed_cols_answer,
            changed_cols_output,
            max_rows_front=10,
            max_rows_back=10,
        )

        prompt_rich = build_prompt_for_case(
            input_info=input_info,
            before_snapshot=before_snapshot_rich,
            query=query_text,
            gt_summary=gt_summary_full,
            generated_text=out_summary_full,
        )
        prompt_rich_chars = len(prompt_rich)

        if prompt_rich_chars <= PROMPT_MAX_CHARS:
            diff_answer_compact = compress_diff(diff_answer_full)
            diff_output_compact = compress_diff(diff_output_full)

            bundle_compact = {
                "input_file": str(input_path),
                "answer_file": str(answer_path),
                "output_file": str(output_path),
                "before": before_compact_rich,
                "diff_answer": diff_answer_compact,
                "diff_output": diff_output_compact,
                "mode": "rich",
                "task_classification": "modify",
            }

            print("=== COMPACT CASE BUNDLE (RICH, modify) ===")

            return {
                "bundle": bundle_compact,
                "before_snapshot": before_snapshot_rich,
                "gt_summary": gt_summary_full,
                "out_summary": out_summary_full,
                "prompt": prompt_rich,
                "prompt_chars": prompt_rich_chars,
                "task_classification": "modify",
            }

        # ---------- ， SIMPLE ---------- 
        before_snapshot_simple = build_before_snapshot_simple(
            before_sheets,
            changed_rows_answer,
            changed_rows_output,
            max_rows_front=10,
            max_rows_back=10,
            max_cols=5,
            max_chars=500000,
        )
        before_compact_simple = build_compressed_before_simple(
            before_sheets,
            changed_rows_answer,
            changed_rows_output,
            max_rows_front=10,
            max_rows_back=10,
            max_cols=5,
        )

        prompt_simple = build_prompt_for_case(
            input_info=input_info,
            before_snapshot=before_snapshot_simple,
            query=query_text,
            gt_summary=gt_summary_full,
            generated_text=out_summary_full,
        )
        prompt_simple_chars = len(prompt_simple)

        diff_answer_compact = compress_diff(diff_answer_full)
        diff_output_compact = compress_diff(diff_output_full)

        bundle_compact_simple = {
            "input_file": str(input_path),
            "answer_file": str(answer_path),
            "output_file": str(output_path),
            "before": before_compact_simple,
            "diff_answer": diff_answer_compact,
            "diff_output": diff_output_compact,
            "mode": "simple",
            "task_classification": "modify",
        }

        print("=== COMPACT CASE BUNDLE (SIMPLE, modify fallback) ===")

        return {
            "bundle": bundle_compact_simple,
            "before_snapshot": before_snapshot_simple,
            "gt_summary": gt_summary_full,
            "out_summary": out_summary_full,
            "prompt": prompt_simple,
            "prompt_chars": prompt_simple_chars,
            "task_classification": "modify",
        }

    # ========= ： answer vs output，（answer/output “”） ========= 
    if file_mode == "generate":
        # case 
        case_dir = input_path.parent

        # input.txt / answer.txt / output.txt（） 
        input_txt = read_optional_text_file(case_dir / "input.txt", label="input.txt")
        answer_txt = read_optional_text_file(case_dir / "answer.txt", label="answer.txt")
        output_txt = read_optional_text_file(case_dir / "output.txt", label="output.txt")
        
        
        # 1) answer / output （） 
        has_answer_xlsx = answer_path.is_file()
        has_output_xlsx = output_path.is_file()

        answer_excel_text = ""
        output_excel_text = ""

        if has_answer_xlsx:
            answer_excel_text = build_snapshot_full(
                answer_path,
                max_chars=PROMPT_MAX_CHARS,
            )

        if has_output_xlsx:
            output_excel_text = build_snapshot_full(
                output_path,
                max_chars=PROMPT_MAX_CHARS,
            )

        # 1b) input/、answer/、output/ “ + ” 
        input_folder = case_dir / "input"
        answer_folder = case_dir / "answer"
        output_folder = case_dir / "output"

        input_rich_text, input_images = load_text_image_sequence_from_folder(input_folder)
        answer_folder_text, answer_images = load_text_image_sequence_from_folder(answer_folder)
        output_folder_text, output_images = load_text_image_sequence_from_folder(output_folder)

        # input.txt， input_rich_text 
        if input_txt:
            extra = "【In addition to other content, the output also includes the following text】\n" + input_txt
            if input_rich_text:
                input_rich_text = input_rich_text + "\n\n" + extra
            else:
                input_rich_text = extra


        answer_full_parts: List[str] = []
        if answer_folder_text:
            answer_full_parts.append("【answer包含的markdown/docx文档转换为的文本和图片序列】\n" + answer_folder_text)
        if answer_excel_text:
            answer_full_parts.append("【answer.xlsx 的完整内容】\n" + answer_excel_text)
        if answer_txt:
            answer_full_parts.append("【In addition to other content, the answer also includes the following text】\n" + answer_txt)
        if answer_full_parts:
            answer_full_text = "\n\n".join(answer_full_parts)
        else:
            answer_full_text = "（本 case 未提供 answer.xlsx 或 answer文本内容，请主要结合截图与其它附加信息进行判断。）"

        output_full_parts: List[str] = []
        if output_folder_text:
            output_full_parts.append("【output包含的markdown/docx文档转换为的文本和图片序列】\n" + output_folder_text)
        if output_excel_text:
            output_full_parts.append("【output.xlsx 的完整内容】\n" + output_excel_text)
        if output_txt:
            output_full_parts.append("【In addition to other content, the answer also includes the following text】\n" + output_txt)
        if output_full_parts:
            output_full_text = "\n\n".join(output_full_parts)
        else:
            output_full_text = "（本 case 未提供 output.xlsx 或 output文本内容，模型的结果可能仅以图片或其它形式给出，请结合这些信息进行判断。）"

        # 2) .xlsx （ answer / output ） 
        extra_xlsx_paths: List[Path] = []
        for p in sorted(case_dir.glob("*.xlsx")):
            # answer / output （） 
            if p.resolve() in (answer_path.resolve(), output_path.resolve()):
                continue
            extra_xlsx_paths.append(p)

        
        extra_workbook_contents: Dict[str, str] = {}
        extra_sections: List[str] = []
        for p in extra_xlsx_paths:
            snapshot = build_snapshot_full(
                p,
                max_chars=PROMPT_MAX_CHARS,
            )
            extra_workbook_contents[p.name] = snapshot
            section = f"### 附加工作簿 {p.name}\n\n{snapshot}\n\n"
            extra_sections.append(section)

        # 3) prompt， 
        base_prompt = JUDGE_PROMPT_GENERATE.format(
            query=query_text,
            input_rich_text=input_rich_text,
            answer_full_text=answer_full_text,
            output_full_text=output_full_text,
            extra_workbooks_text="",
        )
        base_len = len(base_prompt)
        remaining_chars = max(0, PROMPT_MAX_CHARS - base_len)

        # 4) ， 
        selected_sections: List[str] = []
        used_len = 0
        for section in extra_sections:
            slen = len(section)
            if used_len + slen > remaining_chars:
                break
            selected_sections.append(section)
            used_len += slen

        if selected_sections:
            extra_workbooks_text = "".join(selected_sections)
        else:
            if extra_xlsx_paths:
                extra_workbooks_text = "（由于长度限制，附加工作簿的详细内容被省略，仅供你知道还存在其它工作簿。）"
            else:
                extra_workbooks_text = "（无附加工作簿。）"

        prompt = JUDGE_PROMPT_GENERATE.format(
            query=query_text,
            input_rich_text=input_rich_text,
            answer_full_text=answer_full_text,
            output_full_text=output_full_text,
            extra_workbooks_text=extra_workbooks_text,
        )
        prompt_chars = len(prompt)

        bundle_compact = {
            "answer_file": str(answer_path),
            "output_file": str(output_path),
            "answer_full_text": answer_full_text,
            "output_full_text": output_full_text,
            "extra_workbooks": extra_workbook_contents,
            "mode": "rich",
            "task_classification": "generate",
            # ： xlsx 
            "has_answer_xlsx": has_answer_xlsx,
            "has_output_xlsx": has_output_xlsx,
            # input/、answer/、output/ （） 
            "folder_images": {
                "input": [str(p.resolve()) for p in input_images],
                "answer": [str(p.resolve()) for p in answer_images],
                "output": [str(p.resolve()) for p in output_images],
            },
        }

        print("=== COMPACT CASE BUNDLE (RICH, generate) ===")
        try:
            print(json.dumps(bundle_compact, ensure_ascii=False))
        except Exception:
            print("[WARN] Failed to dump compact bundle (generate) as JSON string.")

        return {
            "bundle": bundle_compact,
            "answer_full_text": answer_full_text,
            "output_full_text": output_full_text,
            "prompt": prompt,
            "prompt_chars": prompt_chars,
            "task_classification": "generate",
        }

    # ========= QA ：Question & Answer， query + answer/output（） ========= 
    if file_mode == "qa":
        case_dir = input_path.parent

        # input.txt / answer.txt / output.txt（） 
        input_txt = read_optional_text_file(case_dir / "input.txt", label="input.txt")
        answer_txt = read_optional_text_file(case_dir / "answer.txt", label="answer.txt")
        output_txt = read_optional_text_file(case_dir / "output.txt", label="output.txt")

        input_folder = case_dir / "input"
        answer_folder = case_dir / "answer"
        output_folder = case_dir / "output"


        input_rich_text, input_images = load_text_image_sequence_from_folder(input_folder)
        answer_rich_text, answer_images = load_text_image_sequence_from_folder(answer_folder)
        output_rich_text, output_images = load_text_image_sequence_from_folder(output_folder)

        # answer.xlsx / output.xlsx， 
        if not answer_rich_text and answer_path.is_file():
            tmp = build_snapshot_full(
                answer_path,
                max_chars=PROMPT_MAX_CHARS,
            )
            answer_rich_text = "【来自 answer.xlsx 的完整内容】\n" + tmp

        if not output_rich_text and output_path.is_file():
            tmp = build_snapshot_full(
                output_path,
                max_chars=PROMPT_MAX_CHARS,
            )
            output_rich_text = "【来自 output.xlsx 的完整内容】\n" + tmp
            
                # ， input.txt / answer.txt / output.txt 
        if input_txt:
            extra = "【来自 input.txt 的文本内容】\n" + input_txt
            if input_rich_text:
                input_rich_text = input_rich_text + "\n\n" + extra
            else:
                input_rich_text = extra

        if answer_txt:
            extra = "【来自 answer.txt 的文本内容】\n" + answer_txt
            if answer_rich_text:
                answer_rich_text = answer_rich_text + "\n\n" + extra
            else:
                answer_rich_text = extra

        if output_txt:
            extra = "【来自 output.txt 的文本内容】\n" + output_txt
            if output_rich_text:
                output_rich_text = output_rich_text + "\n\n" + extra
            else:
                output_rich_text = extra


        prompt = JUDGE_PROMPT_QA.format(
            query=query_text,
            input_rich_text=input_rich_text,
            answer_rich_text=answer_rich_text,
            output_rich_text=output_rich_text,
        )
        prompt_chars = len(prompt)

        bundle_compact = {
            "answer_file": str(answer_path),
            "output_file": str(output_path),
            "input_folder_text": input_rich_text,
            "answer_folder_text": answer_rich_text,
            "output_folder_text": output_rich_text,
            "mode": "qa",
            "task_classification": "qa",
            "folder_images": {
                "input": [str(p.resolve()) for p in input_images],
                "answer": [str(p.resolve()) for p in answer_images],
                "output": [str(p.resolve()) for p in output_images],
            },
        }

        print("=== COMPACT CASE BUNDLE (QA) ===")
        try:
            print(json.dumps(bundle_compact, ensure_ascii=False))
        except Exception:
            print("[WARN] Failed to dump compact bundle (qa) as JSON string.")

        return {
            "bundle": bundle_compact,
            "prompt": prompt,
            "prompt_chars": prompt_chars,
            "task_classification": "qa",
        }
