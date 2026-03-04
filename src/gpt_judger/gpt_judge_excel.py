# gpt_judge_excel.py
"""
Helpers for working with Excel workbooks:
- A1 address parsing
- BEFORE snapshots (rich/simple/GPT-selected)
- Full snapshot (value + formula)
- Diff compression and summarization
"""

from pathlib import Path
from typing import Dict, Any, List, Optional, Tuple, Set

import re
from openpyxl import load_workbook

_ADDR_RE = re.compile(r"^([A-Za-z]+)(\d+)$")


# ---------- Address helpers ----------

def _col_letters_to_index(col_letters: str) -> int:
    """Convert Excel column letters to 1-based index (A -> 1, AA -> 27)."""
    col_idx = 0
    for ch in col_letters.upper():
        col_idx = col_idx * 26 + (ord(ch) - ord("A") + 1)
    return col_idx


def _index_to_col_letters(idx: int) -> str:
    """Convert 1-based index to Excel column letters (1 -> A, 27 -> AA)."""
    letters: List[str] = []
    while idx > 0:
        idx, rem = divmod(idx - 1, 26)
        letters.append(chr(ord("A") + rem))
    return "".join(reversed(letters))


def parse_addr(addr: str) -> Optional[Tuple[int, int]]:
    """
    Parse A1-style addresses into (row, col_idx), col_idx is 1-based.
    Returns None if parsing fails.
    """
    m = _ADDR_RE.match(addr)
    if not m:
        return None
    col_letters, row_str = m.groups()
    row = int(row_str)
    col_idx = _col_letters_to_index(col_letters)
    return row, col_idx


# ---------- Diff row / col helpers ----------

def collect_changed_rows(diff_full: Dict[str, Any]) -> Dict[str, Set[int]]:
    """Collect changed row indices per sheet from a full diff JSON."""
    result: Dict[str, Set[int]] = {}
    sheets_diff = diff_full.get("sheets_diff", {}) or {}
    for sheet_name, sheet_info in sheets_diff.items():
        diffs = sheet_info.get("diff", []) or []
        for change in diffs:
            addr = change.get("address", "")
            if not addr:
                continue
            parsed = parse_addr(addr)
            if not parsed:
                continue
            row, _ = parsed
            result.setdefault(sheet_name, set()).add(row)
    return result


def collect_changed_cols(diff_full: Dict[str, Any]) -> Dict[str, Set[int]]:
    """Collect changed column indices per sheet from a full diff JSON."""
    result: Dict[str, Set[int]] = {}
    sheets_diff = diff_full.get("sheets_diff", {}) or {}
    for sheet_name, sheet_info in sheets_diff.items():
        diffs = sheet_info.get("diff", []) or []
        for change in diffs:
            addr = change.get("address", "")
            if not addr:
                continue
            parsed = parse_addr(addr)
            if not parsed:
                continue
            _, col = parsed
            result.setdefault(sheet_name, set()).add(col)
    return result


def pick_rows_to_keep(
    all_rows: List[int],
    changed_rows_answer: Set[int],
    changed_rows_output: Set[int],
    max_rows_front: int,
    max_rows_back: int,
) -> Tuple[List[int], int]:
    """
    Decide which rows to keep for BEFORE snapshot:
      - always keep first `max_rows_front` rows,
      - always keep last `max_rows_back` rows,
      - always keep changed rows.
    Returns (sorted_rows_to_keep, omitted_count).
    """
    if not all_rows:
        return [], 0

    base = set(all_rows[:max_rows_front] + all_rows[-max_rows_back:])
    changed = (changed_rows_answer or set()) | (changed_rows_output or set())
    rows_to_keep = base | (changed & set(all_rows))

    rows_sorted = sorted(rows_to_keep)
    omitted = len(all_rows) - len(rows_sorted)
    return rows_sorted, omitted


# ---------- BEFORE snapshot builders ----------

def build_before_snapshot_rich(
    before_sheets: Dict[str, List[Dict[str, Any]]],
    changed_rows_answer: Dict[str, Set[int]],
    changed_rows_output: Dict[str, Set[int]],
    changed_cols_answer: Dict[str, Set[int]],
    changed_cols_output: Dict[str, Set[int]],
    max_rows_front: int = 10,
    max_rows_back: int = 10,
    max_chars: int = 500000,
) -> str:
    lines: List[str] = []
    sheet_names = sorted(before_sheets.keys())

    lines.append(
        "=== BEFORE snapshot (RICH) from input.xlsx: first/last rows + all changed rows/columns; all columns ==="
    )

    for sheet_name in sheet_names:
        items = before_sheets.get(sheet_name, []) or []
        # row -> col_idx -> (addr,text,formula)
        row_map: Dict[int, Dict[int, Tuple[str, str, str]]] = {}

        for item in items:
            addr = item.get("address", "")
            if not addr:
                continue
            parsed = parse_addr(addr)
            if not parsed:
                continue
            row, col_idx = parsed

            text = str(item.get("text", "") or "")
            formula = str(item.get("formula", "") or "")
            row_map.setdefault(row, {})[col_idx] = (addr, text, formula)

        if not row_map:
            continue

        all_rows = sorted(row_map.keys())
        total_rows = len(all_rows)

        changed_a_rows = changed_rows_answer.get(sheet_name, set())
        changed_o_rows = changed_rows_output.get(sheet_name, set())
        changed_rows = (changed_a_rows | changed_o_rows) & set(all_rows)

        changed_a_cols = changed_cols_answer.get(sheet_name, set())
        changed_o_cols = changed_cols_output.get(sheet_name, set())
        changed_cols = changed_a_cols | changed_o_cols

        # “/ max + ” 
        selected_rows, omitted_base = pick_rows_to_keep(
            all_rows, changed_a_rows, changed_o_rows, max_rows_front, max_rows_back
        )
        selected_set = set(selected_rows)

        
        for r in all_rows:
            cols = row_map.get(r, {})
            if any(c in changed_cols for c in cols.keys()):
                selected_set.add(r)

        selected_rows = sorted(selected_set)
        omitted_rows = len(all_rows) - len(selected_rows)

        lines.append(
            f"[Sheet {sheet_name}] total rows with data: {total_rows}"
        )
        if omitted_rows > 0:
            lines.append(
                f"... ({omitted_rows} rows omitted in this sheet BEFORE snapshot) ..."
            )

        for r in selected_rows:
            cols = row_map.get(r, {})
            if not cols:
                continue
            cell_strs: List[str] = []
            for c in sorted(cols.keys()):
                addr, text, formula = cols[c]
                if formula:
                    cell_strs.append(
                        f"{addr}(text='{text}', formula='{formula}')"
                    )
                else:
                    cell_strs.append(
                        f"{addr}(text='{text}')"
                    )
            if cell_strs:
                lines.append(
                    f"{sheet_name} Row {r}: " + "; ".join(cell_strs)
                )

    text = "\n".join(lines)
    if len(text) <= max_chars:
        return text
    return text[:max_chars] + "\n...[BEFORE snapshot (RICH) truncated due to length limit]"


def build_before_snapshot_simple(
    before_sheets: Dict[str, List[Dict[str, Any]]],
    changed_rows_answer: Dict[str, Set[int]],
    changed_rows_output: Dict[str, Set[int]],
    max_rows_front: int = 10,
    max_rows_back: int = 10,
    max_cols: int = 5,
    max_chars: int = 500000,
) -> str:
    lines: List[str] = []
    sheet_names = sorted(before_sheets.keys())

    lines.append(
        "=== BEFORE snapshot (SIMPLE) from input.xlsx: first/last rows + changed rows; first 5 columns A–E ==="
    )

    for sheet_name in sheet_names:
        items = before_sheets.get(sheet_name, []) or []
        row_map: Dict[int, Dict[int, Tuple[str, str, str]]] = {}

        for item in items:
            addr = item.get("address", "")
            if not addr:
                continue
            parsed = parse_addr(addr)
            if not parsed:
                continue
            row, col_idx = parsed
            if col_idx > max_cols:
                continue

            text = str(item.get("text", "") or "")
            formula = str(item.get("formula", "") or "")
            row_map.setdefault(row, {})[col_idx] = (addr, text, formula)

        if not row_map:
            continue

        all_rows = sorted(row_map.keys())
        total_rows = len(all_rows)

        changed_a = changed_rows_answer.get(sheet_name, set())
        changed_o = changed_rows_output.get(sheet_name, set())
        selected_rows, omitted_rows = pick_rows_to_keep(
            all_rows, changed_a, changed_o, max_rows_front, max_rows_back
        )

        lines.append(
            f"[Sheet {sheet_name}] total rows with data (first {max_cols} cols): {total_rows}"
        )
        if omitted_rows > 0:
            lines.append(
                f"... ({omitted_rows} rows omitted in this sheet BEFORE snapshot) ..."
            )

        for r in selected_rows:
            cols = row_map.get(r, {})
            cell_strs: List[str] = []
            for c in range(1, max_cols + 1):
                if c not in cols:
                    continue
                addr, text, formula = cols[c]
                if formula:
                    cell_strs.append(
                        f"{addr}(text='{text}', formula='{formula}')"
                    )
                else:
                    cell_strs.append(
                        f"{addr}(text='{text}')"
                    )
            if cell_strs:
                lines.append(
                    f"{sheet_name} Row {r}: " + "; ".join(cell_strs)
                )

    text = "\n".join(lines)
    if len(text) <= max_chars:
        return text
    return text[:max_chars] + "\n...[BEFORE snapshot (SIMPLE) truncated due to length limit]"


def build_compressed_before_rich(
    before_sheets: Dict[str, List[Dict[str, Any]]],
    changed_rows_answer: Dict[str, Set[int]],
    changed_rows_output: Dict[str, Set[int]],
    changed_cols_answer: Dict[str, Set[int]],
    changed_cols_output: Dict[str, Set[int]],
    max_rows_front: int = 10,
    max_rows_back: int = 10,
) -> Dict[str, Dict[str, Dict[str, Dict[str, str]]]]:
    result: Dict[str, Dict[str, Dict[str, Dict[str, str]]]] = {}

    for sheet_name, items in before_sheets.items():
        row_map: Dict[int, Dict[int, Tuple[str, str, str]]] = {}
        for item in items or []:
            addr = item.get("address", "")
            if not addr:
                continue
            parsed = parse_addr(addr)
            if not parsed:
                continue
            row, col_idx = parsed
            text = str(item.get("text", "") or "")
            formula = str(item.get("formula", "") or "")
            row_map.setdefault(row, {})[col_idx] = (addr, text, formula)

        if not row_map:
            continue

        all_rows = sorted(row_map.keys())

        changed_a_rows = changed_rows_answer.get(sheet_name, set())
        changed_o_rows = changed_rows_output.get(sheet_name, set())
        changed_a_cols = changed_cols_answer.get(sheet_name, set())
        changed_o_cols = changed_cols_output.get(sheet_name, set())
        changed_cols = changed_a_cols | changed_o_cols

        selected_rows, _ = pick_rows_to_keep(
            all_rows, changed_a_rows, changed_o_rows, max_rows_front, max_rows_back
        )
        selected_set = set(selected_rows)

        
        for r in all_rows:
            cols = row_map.get(r, {})
            if any(c in changed_cols for c in cols.keys()):
                selected_set.add(r)

        selected_rows = sorted(selected_set)
        sheet_obj: Dict[str, Dict[str, Dict[str, str]]] = {}

        for r in selected_rows:
            cols = row_map.get(r, {})
            if not cols:
                continue
            row_key = str(r)
            row_obj: Dict[str, Dict[str, str]] = {}
            for c, (_, text, formula) in cols.items():
                col_letter = _index_to_col_letters(c)
                cell_obj: Dict[str, str] = {}
                if text:
                    cell_obj["text"] = text
                if formula:
                    cell_obj["formula"] = formula
                if cell_obj:
                    row_obj[col_letter] = cell_obj
            if row_obj:
                sheet_obj[row_key] = row_obj

        if sheet_obj:
            result[sheet_name] = sheet_obj

    return result


def build_compressed_before_simple(
    before_sheets: Dict[str, List[Dict[str, Any]]],
    changed_rows_answer: Dict[str, Set[int]],
    changed_rows_output: Dict[str, Set[int]],
    max_rows_front: int = 10,
    max_rows_back: int = 10,
    max_cols: int = 5,
) -> Dict[str, Dict[str, Dict[str, Dict[str, str]]]]:
    result: Dict[str, Dict[str, Dict[str, Dict[str, str]]]] = {}

    for sheet_name, items in before_sheets.items():
        row_map: Dict[int, Dict[int, Tuple[str, str, str]]] = {}
        for item in items or []:
            addr = item.get("address", "")
            if not addr:
                continue
            parsed = parse_addr(addr)
            if not parsed:
                continue
            row, col_idx = parsed
            if col_idx > max_cols:
                continue
            text = str(item.get("text", "") or "")
            formula = str(item.get("formula", "") or "")
            row_map.setdefault(row, {})[col_idx] = (addr, text, formula)

        if not row_map:
            continue

        all_rows = sorted(row_map.keys())
        changed_a = changed_rows_answer.get(sheet_name, set())
        changed_o = changed_rows_output.get(sheet_name, set())
        selected_rows, _ = pick_rows_to_keep(
            all_rows, changed_a, changed_o, max_rows_front, max_rows_back
        )

        sheet_obj: Dict[str, Dict[str, Dict[str, str]]] = {}
        for r in selected_rows:
            cols = row_map.get(r, {})
            if not cols:
                continue
            row_key = str(r)
            row_obj: Dict[str, Dict[str, str]] = {}
            for c in range(1, max_cols + 1):
                if c not in cols:
                    continue
                _, text, formula = cols[c]
                col_letter = _index_to_col_letters(c)
                cell_obj: Dict[str, str] = {}
                if text:
                    cell_obj["text"] = text
                if formula:
                    cell_obj["formula"] = formula
                if cell_obj:
                    row_obj[col_letter] = cell_obj
            if row_obj:
                sheet_obj[row_key] = row_obj

        if sheet_obj:
            result[sheet_name] = sheet_obj

    return result

def build_before_snapshot_gpt_cells(
    before_sheets: Dict[str, List[Dict[str, Any]]],
    important_cells: Dict[str, Set[str]],
    max_chars: int = 500000,
) -> str:
    lines: List[str] = []
    sheet_names = sorted(important_cells.keys())

    lines.append(
        "=== BEFORE snapshot (GPT-selected regions only) from input.xlsx ==="
    )

    for sheet_name in sheet_names:
        addr_set = important_cells.get(sheet_name, set())
        if not addr_set:
            continue

        # row -> col_idx -> (addr,text,formula)
        row_map: Dict[int, Dict[int, Tuple[str, str, str]]] = {}

        items = before_sheets.get(sheet_name, []) or []
        for item in items:
            addr = item.get("address", "")
            if not addr or addr not in addr_set:
                continue
            parsed = parse_addr(addr)
            if not parsed:
                continue
            row, col_idx = parsed
            text = str(item.get("text", "") or "")
            formula = str(item.get("formula", "") or "")
            row_map.setdefault(row, {})[col_idx] = (addr, text, formula)

        if not row_map:
            continue

        all_rows = sorted(row_map.keys())
        lines.append(f"[Sheet {sheet_name}] rows in GPT-selected regions: {len(all_rows)}")

        for r in all_rows:
            cols = row_map.get(r, {})
            if not cols:
                continue
            cell_strs: List[str] = []
            for c in sorted(cols.keys()):
                addr, text, formula = cols[c]
                if formula:
                    cell_strs.append(f"{addr}(text='{text}', formula='{formula}')")
                else:
                    cell_strs.append(f"{addr}(text='{text}')")
            if cell_strs:
                lines.append(f"{sheet_name} Row {r}: " + "; ".join(cell_strs))

    text = "\n".join(lines)
    if len(text) <= max_chars:
        return text
    return text[:max_chars] + "\n...[BEFORE snapshot (GPT) truncated due to length limit]"

def build_compressed_before_gpt_cells(
    before_sheets: Dict[str, List[Dict[str, Any]]],
    important_cells: Dict[str, Set[str]],
) -> Dict[str, Dict[str, Dict[str, Dict[str, str]]]]:
    result: Dict[str, Dict[str, Dict[str, Dict[str, str]]]] = {}

    for sheet_name, addr_set in important_cells.items():
        if not addr_set:
            continue

        sheet_items = before_sheets.get(sheet_name, []) or []
        row_map: Dict[int, Dict[int, Tuple[str, str, str]]] = {}

        for item in sheet_items:
            addr = item.get("address", "")
            if not addr or addr not in addr_set:
                continue
            parsed = parse_addr(addr)
            if not parsed:
                continue
            row, col_idx = parsed
            text = str(item.get("text", "") or "")
            formula = str(item.get("formula", "") or "")
            row_map.setdefault(row, {})[col_idx] = (addr, text, formula)

        if not row_map:
            continue

        sheet_obj: Dict[str, Dict[str, Dict[str, str]]] = {}
        for r in sorted(row_map.keys()):
            cols = row_map[r]
            if not cols:
                continue
            row_key = str(r)
            row_obj: Dict[str, Dict[str, str]] = {}
            for c, (_, text, formula) in sorted(cols.items()):
                col_letter = _index_to_col_letters(c)
                cell_obj: Dict[str, str] = {}
                if text:
                    cell_obj["text"] = text
                if formula:
                    cell_obj["formula"] = formula
                if cell_obj:
                    row_obj[col_letter] = cell_obj
            if row_obj:
                sheet_obj[row_key] = row_obj

        if sheet_obj:
            result[sheet_name] = sheet_obj

    return result


# ---------- Workbook snapshot (full) ----------

def build_snapshot_full(
    workbook_path: Path,
    max_chars: int = 500000,
) -> str:
    # （data_only=False），（data_only=True） 
    wb_formula = load_workbook(filename=str(workbook_path), data_only=False)
    wb_values = load_workbook(filename=str(workbook_path), data_only=True)

    lines: List[str] = []

    
    for ws_formula in wb_formula.worksheets:
        sheet_name = ws_formula.title
        ws_values = wb_values[sheet_name]

        lines.append(f"[Sheet {sheet_name}]")

        for row in ws_formula.iter_rows():
            for cell_f in row:
                # ✅ coordinate， MergedCell column_letter 
                addr = cell_f.coordinate

                cell_v = ws_values[addr]

                
                formula = ""
                if isinstance(cell_f.value, str) and cell_f.value.startswith("="):
                    formula = cell_f.value

                # （ Excel ） 
                value = cell_v.value

                
                if value is None and not formula:
                    continue

                value_str = "" if value is None else str(value)

                if formula:
                    lines.append(
                        f"{sheet_name} {addr}(value='{value_str}', formula='{formula}')"
                    )
                else:
                    lines.append(
                        f"{sheet_name} {addr}(value='{value_str}')"
                    )

    text = "\n".join(lines)
    if len(text) <= max_chars:
        return text
    return text[:max_chars] + "\n...[OUTPUT snapshot truncated due to length limit]"

IMAGE_EXTS = {".png", ".jpg", ".jpeg", ".bmp"}

# ---------- Diff compression ----------

def compress_diff(diff_full: Dict[str, Any]) -> Dict[str, Dict[str, Dict[str, str]]]:
    result: Dict[str, Dict[str, Dict[str, str]]] = {}
    sheets_diff = diff_full.get("sheets_diff", {}) or {}

    for sheet_name, sheet_info in sheets_diff.items():
        diffs = sheet_info.get("diff", []) or []
        sheet_map: Dict[str, Dict[str, str]] = {}
        for change in diffs:
            addr = change.get("address", "")
            if not addr:
                continue
            tb = str(change.get("text_before", "") or "")
            ta = str(change.get("text_after", "") or "")
            fb = str(change.get("formula_before", "") or "")
            fa = str(change.get("formula_after", "") or "")

            fields: Dict[str, str] = {}
            if tb != ta:
                fields["text_before"] = tb
                fields["text_after"] = ta
            if fb != fa:
                fields["formula_before"] = fb
                fields["formula_after"] = fa

            if fields:
                sheet_map[addr] = fields

        if sheet_map:
            result[sheet_name] = sheet_map

    return result

def summarize_diff_only(
    diff_full: Dict[str, Any],
    max_chars: int = 500000,
) -> str:
    if not diff_full:
        return "No diff JSON is available."

    sheets_diff = diff_full.get("sheets_diff", {}) or {}
    changed_sheets = sorted(sheets_diff.keys())

    lines: List[str] = []
    lines.append("=== DIFF summary (input → target) ===")
    lines.append(f"Changed sheets count: {len(changed_sheets)}")

    for sheet_name in changed_sheets:
        sheet_info = sheets_diff.get(sheet_name, {}) or {}
        diffs = sheet_info.get("diff", []) or []
        lines.append(f"[Sheet {sheet_name}] diff cells: {len(diffs)}")

        for change in diffs:
            addr = change.get("address", "")
            tb = str(change.get("text_before", "") or "")
            ta = str(change.get("text_after", "") or "")
            fb = str(change.get("formula_before", "") or "")
            fa = str(change.get("formula_after", "") or "")

            parts = [f"text: '{tb}' -> '{ta}'"]
            if fb or fa:
                parts.append(f"formula: '{fb}' -> '{fa}'")
            lines.append(f"{sheet_name} {addr}: " + "; ".join(parts))

    text = "\n".join(lines)
    if len(text) <= max_chars:
        return text
    return text[:max_chars] + "\n...[DIFF summary truncated due to length limit]"


def expand_cell_range(start_addr: str, end_addr: str) -> Set[str]:
    parsed_start = parse_addr(start_addr)
    parsed_end = parse_addr(end_addr)
    if not parsed_start or not parsed_end:
        return set()

    r1, c1 = parsed_start
    r2, c2 = parsed_end
    if r1 > r2:
        r1, r2 = r2, r1
    if c1 > c2:
        c1, c2 = c2, c1

    cells: Set[str] = set()
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            col_letters = _index_to_col_letters(c)
            cells.add(f"{col_letters}{r}")
    return cells

def build_sheet_cells_text_for_region(
    sheet_items: List[Dict[str, Any]],
    max_chars: int = 8000,
) -> str:
    rows: List[Tuple[int, int, str]] = []
    for item in sheet_items or []:
        addr = item.get("address", "")
        if not addr:
            continue
        parsed = parse_addr(addr)
        if not parsed:
            continue
        row, col_idx = parsed
        text = str(item.get("text", "") or "")
        formula = str(item.get("formula", "") or "")
        if formula:
            desc = f"{addr}(text='{text}', formula='{formula}')"
        else:
            desc = f"{addr}(text='{text}')"
        rows.append((row, col_idx, desc))

    rows.sort(key=lambda x: (x[0], x[1]))
    lines = [desc for _, _, desc in rows]
    s = "\n".join(lines)
    if len(s) > max_chars:
        s = s[:max_chars] + "\n...[sheet cells truncated]"
    return s