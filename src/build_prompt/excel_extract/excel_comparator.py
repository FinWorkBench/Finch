"""
Excel File Comparator
Compare differences between two Excel files (cell values, formulas, charts, formatting)

Usage:
    python excel_comparator.py file1.xlsx file2.xlsx --output diff.json
    
Requirements:
    pip install openpyxl
"""

import argparse
import json
import re
from pathlib import Path
from typing import Dict, List, Tuple, Optional, Set

import openpyxl
from openpyxl.chartsheet import Chartsheet


# ==================== Utility Functions ====================

def addr_key(addr: str):
    """Provide sorting key for cell addresses, e.g., A1, B10, AA3"""
    pattern = re.compile(r"^([A-Za-z]+)(\d+)$")
    match = pattern.match(addr)
    if not match:
        return (10**9, 10**9)
    
    col_letters, row_str = match.groups()
    row = int(row_str)
    
    # Convert column letters to numeric index
    col_idx = 0
    for ch in col_letters.upper():
        col_idx = col_idx * 26 + (ord(ch) - ord("A") + 1)
    
    return (row, col_idx)


# ==================== Read Worksheet Content ====================

def read_sheet_to_maps(ws_formula, ws_values=None) -> Tuple[Dict[str, str], Dict[str, str], List[dict]]:
    """
    Read a single worksheet and return:
      - addr2text: {address -> display text/calculated result}
      - addr2form: {address -> formula}
      - items: sorted list of cells [{"address", "text", "formula"}]
    
    Parameters:
    - ws_formula: sheet opened with data_only=False, for reading formulas
    - ws_values: sheet opened with data_only=True, for reading display values/calculated results
    """
    addr2text: Dict[str, str] = {}
    addr2form: Dict[str, str] = {}
    
    for row in ws_formula.iter_rows(values_only=False):
        for cell in row:
            addr = cell.coordinate
            
            # Original value (may be a formula)
            f_val = cell.value
            
            # Display value/calculated result: prefer getting from ws_values
            if ws_values is not None:
                v_cell = ws_values[addr]
                v_val = v_cell.value
            else:
                v_val = cell.value
            
            if f_val is None and v_val is None:
                continue
            
            # Record formulas separately
            if isinstance(f_val, str) and f_val.startswith("="):
                addr2form[addr] = f_val
            
            # Record display values
            if v_val is not None:
                text = str(v_val).strip()
                if text:
                    addr2text[addr] = text
    
    # Build sorted list of cells
    items: List[dict] = []
    all_addrs = sorted(set(addr2text.keys()) | set(addr2form.keys()), key=addr_key)
    
    for addr in all_addrs:
        item = {"address": addr, "text": addr2text.get(addr, "")}
        if addr in addr2form:
            item["formula"] = addr2form[addr]
        items.append(item)
    
    return addr2text, addr2form, items


def read_workbook(path: Path) -> Dict[str, Tuple]:
    """
    Read Excel workbook and return content for each worksheet:
    {sheet_name -> (addr2text, addr2form, items)}
    """
    path = Path(path).resolve()

    # For reading formulas
    wb_formula = openpyxl.load_workbook(path, data_only=False)
    # For reading display values/calculated results
    wb_values = openpyxl.load_workbook(path, data_only=True)

    try:
        # Check and skip chart sheets
        for ws in getattr(wb_formula, "_sheets", []):
            if isinstance(ws, Chartsheet):
                print(f"[INFO] Skipping chart sheet: '{ws.title}' (chart sheets have no cell grid)")

        sheet_maps = {}
        for ws_formula in wb_formula.worksheets:
            title = ws_formula.title
            ws_values_sheet = wb_values[title] if title in wb_values.sheetnames else None

            addr2text, addr2form, items = read_sheet_to_maps(ws_formula, ws_values_sheet)
            sheet_maps[title] = (addr2text, addr2form, items)

        return sheet_maps
    finally:
        try:
            wb_formula.close()
        except Exception:
            pass
        try:
            wb_values.close()
        except Exception:
            pass


# ==================== Comparison Logic ====================

def diff_two_sheets(before_maps, after_maps) -> Tuple[List[dict], List[dict], List[dict]]:
    """
    Compare two worksheets and return:
    - before_items: cell list before modification
    - after_items: cell list after modification
    - diff_items: list of cells with differences
    """
    b_text, b_form, before_items = before_maps
    a_text, a_form, after_items = after_maps
    
    all_addrs = set(b_text.keys()) | set(a_text.keys()) | set(b_form.keys()) | set(a_form.keys())
    diff_items: List[dict] = []
    
    for addr in sorted(all_addrs, key=addr_key):
        tb = b_text.get(addr, "")
        ta = a_text.get(addr, "")
        fb = b_form.get(addr, "")
        fa = a_form.get(addr, "")
        
        # Text comparison (including numeric approximation)
        same_text = (tb == ta)
        
        if not same_text:
            # Try comparing as floating point numbers, keeping 6 decimal places
            try:
                tb_float = float(tb)
                ta_float = float(ta)
                if round(tb_float, 6) == round(ta_float, 6):
                    same_text = True
            except ValueError:
                pass
        
        # Formula comparison (exact string)
        same_formula = (fb == fa)
        
        # Both text and formula are the same => don't record diff
        if same_text and same_formula:
            continue
        
        # Any difference => record diff
        diff_items.append({
            "address": addr,
            "text_before": tb,
            "text_after": ta,
            "formula_before": fb,
            "formula_after": fa,
        })
    
    return before_items, after_items, diff_items


def build_workbook_diff(before_path: Path, after_path: Path) -> dict:
    """
    Build complete difference comparison between two workbooks
    
    Return format:
    {
        "before_file": file path,
        "after_file": file path,
        "before_sheets": {sheet_name: [cell list]},
        "after_sheets": {sheet_name: [cell list]},
        "sheets_diff": {sheet_name: {"before": [], "after": [], "diff": []}}
    }
    """
    before_wb = read_workbook(before_path)
    after_wb = read_workbook(after_path)
    
    all_sheets = set(before_wb.keys()) | set(after_wb.keys())
    
    result = {
        "before_file": str(before_path),
        "after_file": str(after_path),
        "before_sheets": {},
        "after_sheets": {},
        "sheets_diff": {},
    }
    
    for sheet in sorted(all_sheets):
        # Get before and after worksheet content
        if sheet in before_wb:
            before_maps = before_wb[sheet]
        else:
            before_maps = ({}, {}, [])
        
        if sheet in after_wb:
            after_maps = after_wb[sheet]
        else:
            after_maps = ({}, {}, [])
        
        before_items, after_items, diff_items = diff_two_sheets(before_maps, after_maps)
        
        # Record complete before/after content
        result["before_sheets"][sheet] = before_items
        result["after_sheets"][sheet] = after_items
        
        # Only put sheets with differences into sheets_diff
        if diff_items or sheet not in before_wb or sheet not in after_wb:
            result["sheets_diff"][sheet] = {
                "before": before_items,
                "after": after_items,
                "diff": diff_items,
            }
    
    return result


# ==================== Chart Comparison ====================

def get_chart_signature(ws) -> List[Dict]:
    """Build chart signature for worksheet comparison"""
    sig_list: List[Dict] = []
    
    for chart in getattr(ws, "_charts", []):
        chart_type = type(chart).__name__
        
        # Get chart title
        title = ""
        try:
            if getattr(chart, "title", None):
                title = str(chart.title)
        except:
            pass
        
        # Get chart position
        position = ""
        try:
            if getattr(chart, "anchor", None):
                position = str(chart.anchor)
        except:
            pass
        
        # Get series information
        series_info: List[Dict] = []
        try:
            if getattr(chart, "series", None):
                for s in chart.series:
                    try:
                        v = getattr(s, "values", None)
                        xv = getattr(s, "xvalues", None)
                        series_info.append({
                            "values": "" if v is None else str(v),
                            "xvalues": "" if xv is None else str(xv),
                        })
                    except:
                        continue
        except:
            pass
        
        sig_list.append({
            "type": chart_type,
            "title": title,
            "position": position,
            "series": series_info,
        })
    
    sig_list.sort(key=lambda c: (c["type"], c["title"], c["position"], str(c["series"])))
    return sig_list


def get_chart_diff_sheets(wb1_path: Path, wb2_path: Path) -> List[str]:
    """Compare chart differences and return list of worksheet names with chart changes"""
    changed: List[str] = []
    
    try:
        wb1 = openpyxl.load_workbook(wb1_path, data_only=False)
        wb2 = openpyxl.load_workbook(wb2_path, data_only=False)
    except Exception as e:
        print(f"[ERROR] Unable to open workbooks for chart comparison: {e}")
        return changed

    try:
        names1 = set(wb1.sheetnames)
        names2 = set(wb2.sheetnames)
        common = names1 & names2

        for name in common:
            ws1 = wb1[name]
            ws2 = wb2[name]

            sig1 = get_chart_signature(ws1)
            sig2 = get_chart_signature(ws2)

            if sig1 != sig2:
                changed.append(name)

        return sorted(changed)
    finally:
        try:
            wb1.close()
        except Exception:
            pass
        try:
            wb2.close()
        except Exception:
            pass


# ==================== Formatting Comparison ====================

def get_cell_style_key(cell) -> tuple:
    """Extract key information from cell's visible style for comparison"""
    b = cell.border
    f = cell.fill
    ft = cell.font
    al = cell.alignment
    
    border_key = (
        b.left.style, getattr(b.left.color, "rgb", None),
        b.right.style, getattr(b.right.color, "rgb", None),
        b.top.style, getattr(b.top.color, "rgb", None),
        b.bottom.style, getattr(b.bottom.color, "rgb", None),
    )
    
    fill_key = (
        f.patternType,
        getattr(f.fgColor, "type", None), getattr(f.fgColor, "rgb", None),
        getattr(f.bgColor, "type", None), getattr(f.bgColor, "rgb", None),
    )
    
    font_key = (
        ft.name, ft.sz, ft.bold, ft.italic, ft.underline,
        getattr(ft.color, "rgb", None),
    )
    
    align_key = (
        al.horizontal, al.vertical, al.wrapText, al.shrinkToFit,
        al.indent, al.textRotation,
    )
    
    numfmt_key = cell.number_format
    
    return border_key, fill_key, font_key, align_key, numfmt_key


def get_formatting_diff_sheets(wb1_path: Path, wb2_path: Path, 
                               max_rows: int = 500, max_cols: int = 100) -> List[str]:
    """
    Compare cell formatting differences and return list of worksheet names with formatting changes
    For performance control, only compare within the first max_rows × max_cols area
    """
    changed: List[str] = []
    
    try:
        wb1 = openpyxl.load_workbook(wb1_path, data_only=False)
        wb2 = openpyxl.load_workbook(wb2_path, data_only=False)
    except Exception as e:
        print(f"[ERROR] Unable to open workbooks for formatting comparison: {e}")
        return changed

    try:
        names1 = set(wb1.sheetnames)
        names2 = set(wb2.sheetnames)
        common = names1 & names2

        for name in common:
            ws1 = wb1[name]
            ws2 = wb2[name]

            # Skip chart sheets
            if isinstance(ws1, Chartsheet) or isinstance(ws2, Chartsheet):
                print(f"[INFO] Skipping formatting comparison for chart sheet: '{name}'")
                continue

            # Only compare cells within specified range
            rows = min(max(ws1.max_row, ws2.max_row), max_rows)
            cols = min(max(ws1.max_column, ws2.max_column), max_cols)

            has_diff = False
            for r in range(1, rows + 1):
                if has_diff:
                    break
                for c in range(1, cols + 1):
                    c1 = ws1.cell(row=r, column=c)
                    c2 = ws2.cell(row=r, column=c)

                    if get_cell_style_key(c1) != get_cell_style_key(c2):
                        has_diff = True
                        break

            if has_diff:
                changed.append(name)

        return sorted(changed)
    finally:
        try:
            wb1.close()
        except Exception:
            pass
        try:
            wb2.close()
        except Exception:
            pass


# ==================== Comprehensive Comparison ====================

def get_diff_info(
    wb1_path: Path,
    wb2_path: Path,
    include_formatting: bool = True,
    include_charts: bool = True
) -> dict:
    """
    Return a merged diff info dict:
    - detailed diff from build_workbook_diff
    - plus summary with sheet-level categories

    summary keys:
      - cell_changed: sheets that exist in BOTH workbooks and have non-empty cell/formula diffs
      - sheet_added: sheets present only in AFTER
      - sheet_removed: sheets present only in BEFORE
      - chart_changed: sheets with chart changes (only for common sheets)
      - formatting_changed: sheets with formatting changes (only for common sheets)
      - all_changed: union of all above categories
    """
    # 1) detailed diff (cells/formulas + sheet existence captured in sheets_diff)
    diff_result = build_workbook_diff(wb1_path, wb2_path)

    # 2) compute added/removed based on sheetnames from diff_result
    before_sheets = set(diff_result.get("before_sheets", {}).keys())
    after_sheets = set(diff_result.get("after_sheets", {}).keys())

    sheet_added = sorted(after_sheets - before_sheets)
    sheet_removed = sorted(before_sheets - after_sheets)
    common_sheets = before_sheets & after_sheets

    # 3) cell_changed: ONLY common sheets with actual cell/formula diffs
    sheets_diff = diff_result.get("sheets_diff", {})
    cell_changed = sorted([
        sheet_name
        for sheet_name, sheet_info in sheets_diff.items()
        if (sheet_name in common_sheets) and bool(sheet_info.get("diff"))
    ])

    # 4) chart/formatting diffs (only meaningful for common sheets)
    chart_changed: List[str] = []
    if include_charts:
        try:
            chart_changed = get_chart_diff_sheets(wb1_path, wb2_path)
        except Exception as e:
            print(f"[ERROR] Chart comparison failed: {e}")
            chart_changed = []

    formatting_changed: List[str] = []
    if include_formatting:
        try:
            formatting_changed = get_formatting_diff_sheets(wb1_path, wb2_path)
        except Exception as e:
            print(f"[ERROR] Formatting comparison failed: {e}")
            formatting_changed = []

    # 5) all_changed includes added/removed too
    all_changed = sorted(
        set(cell_changed)
        | set(chart_changed)
        | set(formatting_changed)
        | set(sheet_added)
        | set(sheet_removed)
    )

    diff_result["summary"] = {
        "cell_changed": cell_changed,
        "sheet_added": sheet_added,
        "sheet_removed": sheet_removed,
        "chart_changed": chart_changed,
        "formatting_changed": formatting_changed,
        "all_changed": all_changed,
    }
    return diff_result




# ==================== Command Line Interface ====================

def main():
    parser = argparse.ArgumentParser(
        description="Compare differences between two Excel files (cell values, formulas, charts, formatting)"
    )
    parser.add_argument("before", type=Path, help="Path to the 'before' Excel file")
    parser.add_argument("after", type=Path, help="Path to the 'after' Excel file")
    parser.add_argument(
        "-o", "--output", type=Path, default=None,
        help="Output JSON file path (default: ./excel_diff/<after_file_name>_diff.json)"
    )
    parser.add_argument(
        "--no-formatting", action="store_true",
        help="Skip cell formatting comparison"
    )
    parser.add_argument(
        "--no-charts", action="store_true",
        help="Skip chart comparison"
    )
    parser.add_argument(
        "--summary-only", action="store_true",
        help="Output only change summary, not detailed differences"
    )
    
    args = parser.parse_args()
    
    before = args.before.resolve()
    after = args.after.resolve()
    
    if not before.exists():
        print(f"[ERROR] File not found: {before}")
        return
    
    if not after.exists():
        print(f"[ERROR] File not found: {after}")
        return
    
    print(f"[INFO] Comparing files...")
    print(f"  Before: {before}")
    print(f"  After:  {after}")
    
    # Execute comparison
    if args.summary_only:
        # Generate summary only
        diff_info = get_diff_info(
            before, after,
            include_formatting=not args.no_formatting,
            include_charts=not args.no_charts
        )
        output_data = {
            "before_file": str(before),
            "after_file": str(after),
            "summary": diff_info["summary"]
        }
    else:
        # Generate complete differences
        output_data = get_diff_info(
            before, after,
            include_formatting=not args.no_formatting,
            include_charts=not args.no_charts
        )
    
    # Determine output path
    if args.output is None:
        out_dir = Path.cwd() / "excel_diff"
        out_dir.mkdir(parents=True, exist_ok=True)
        out_path = out_dir / f"{after.stem}_diff.json"
    else:
        out_path = args.output
        out_path.parent.mkdir(parents=True, exist_ok=True)
    
    # Save results
    with out_path.open("w", encoding="utf-8") as f:
        json.dump(output_data, f, ensure_ascii=False, indent=2)
    
    print(f"\n[SUCCESS] Comparison complete! Differences saved to: {out_path}")
    
    # Print summary
    if "summary" in output_data:
        summary = output_data["summary"]
        print(f"\nChange Summary:")
        print(f"  Cell content changes: {len(summary['cell_changed'])} worksheets")
        if summary['cell_changed']:
            print(f"    {', '.join(summary['cell_changed'])}")
        
        if not args.no_charts:
            print(f"  Chart changes: {len(summary['chart_changed'])} worksheets")
            if summary['chart_changed']:
                print(f"    {', '.join(summary['chart_changed'])}")
        
        if not args.no_formatting:
            print(f"  Formatting changes: {len(summary['formatting_changed'])} worksheets")
            if summary['formatting_changed']:
                print(f"    {', '.join(summary['formatting_changed'])}")
        
        print(f"  Total changes: {len(summary['all_changed'])} worksheets")


if __name__ == "__main__":
    main()
