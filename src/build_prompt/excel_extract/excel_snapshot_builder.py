"""
Excel Snapshot Builder
Build text-based snapshots of Excel workbooks for documentation, comparison, or analysis.

Features:
- Full snapshot: All cells with values and formulas
- Rich snapshot: Selective rows/columns with change tracking
- Simple snapshot: Limited columns for quick overview
- JSON export: Structured data export

Usage:
    python excel_snapshot_builder.py input.xlsx --mode full --output snapshot.txt
    python excel_snapshot_builder.py input.xlsx --mode json --output snapshot.json
    
Requirements:
    pip install openpyxl
"""

import argparse
import json
import re
from pathlib import Path
from typing import Dict, Any, List, Optional, Tuple, Set

from openpyxl import load_workbook


# ==================== Address Parsing ====================

_ADDR_RE = re.compile(r"^([A-Za-z]+)(\d+)$")


def col_letters_to_index(col_letters: str) -> int:
    """Convert Excel column letters to 1-based index (A -> 1, AA -> 27)."""
    col_idx = 0
    for ch in col_letters.upper():
        col_idx = col_idx * 26 + (ord(ch) - ord("A") + 1)
    return col_idx


def index_to_col_letters(idx: int) -> str:
    """Convert 1-based index to Excel column letters (1 -> A, 27 -> AA)."""
    letters: List[str] = []
    while idx > 0:
        idx, rem = divmod(idx - 1, 26)
        letters.append(chr(ord("A") + rem))
    return "".join(reversed(letters))


def parse_addr(addr: str) -> Optional[Tuple[int, int]]:
    """
    Parse A1-style addresses into (row, col_idx), col_idx is 1-based.
    Returns None if parsing fails.
    
    Examples:
        parse_addr("A1") -> (1, 1)
        parse_addr("Z10") -> (10, 26)
        parse_addr("AA5") -> (5, 27)
    """
    m = _ADDR_RE.match(addr)
    if not m:
        return None
    col_letters, row_str = m.groups()
    row = int(row_str)
    col_idx = col_letters_to_index(col_letters)
    return row, col_idx


# ==================== Range Expansion ====================

def expand_cell_range(start_addr: str, end_addr: str) -> Set[str]:
    """
    Expand a rectangular range like A1:G5 into a set of all cell addresses.
    
    Examples:
        expand_cell_range("A1", "B2") -> {"A1", "A2", "B1", "B2"}
    """
    parsed_start = parse_addr(start_addr)
    parsed_end = parse_addr(end_addr)
    if not parsed_start or not parsed_end:
        return set()

    r1, c1 = parsed_start
    r2, c2 = parsed_end
    if r1 > r2:
        r1, r2 = r2, r1
    if c1 > c2:
        c1, c2 = c2, c1

    cells: Set[str] = set()
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            col_letters = index_to_col_letters(c)
            cells.add(f"{col_letters}{r}")
    return cells


# ==================== Workbook Reading ====================

def read_workbook_data(workbook_path: Path) -> Dict[str, List[Dict[str, Any]]]:
    """
    Read an Excel workbook and extract all cell data.
    
    Returns:
        {
            "Sheet1": [
                {"address": "A1", "text": "Hello", "formula": ""},
                {"address": "B2", "text": "100", "formula": "=SUM(A1:A10)"},
                ...
            ],
            ...
        }
    """
    # Load workbook twice: once for formulas, once for values
    wb_formula = load_workbook(filename=str(workbook_path), data_only=False)
    wb_values = load_workbook(filename=str(workbook_path), data_only=True)

    try:
        result: Dict[str, List[Dict[str, Any]]] = {}

        for ws_formula in wb_formula.worksheets:
            sheet_name = ws_formula.title
            ws_values_sheet = wb_values[sheet_name]

            items: List[Dict[str, Any]] = []

            for row in ws_formula.iter_rows():
                for cell_f in row:
                    # Use coordinate to avoid issues with merged cells
                    addr = cell_f.coordinate
                    cell_v = ws_values_sheet[addr]

                    # Extract formula if present
                    formula = ""
                    if isinstance(cell_f.value, str) and cell_f.value.startswith("="):
                        formula = cell_f.value

                    # Extract calculated value
                    value = cell_v.value

                    # Skip empty cells
                    if value is None and not formula:
                        continue

                    text = "" if value is None else str(value)

                    items.append({
                        "address": addr,
                        "text": text,
                        "formula": formula
                    })

            if items:
                result[sheet_name] = items

        return result
    finally:
        try:
            wb_formula.close()
        except Exception:
            pass
        try:
            wb_values.close()
        except Exception:
            pass


# ==================== Full Snapshot Builder ====================

def build_snapshot_full(
    workbook_path: Path,
    max_chars: int = 500000,
) -> str:
    """
    Build a complete text snapshot of an Excel workbook.
    Includes all non-empty cells with their values and formulas.
    
    Format:
        [Sheet Sheet1]
        Sheet1 A1(value='100', formula='=SUM(A2:A10)')
        Sheet1 B1(value='hello')
        ...
    
    Args:
        workbook_path: Path to the Excel file
        max_chars: Maximum characters for output (truncates if exceeded)
    
    Returns:
        Text snapshot of the workbook
    """
    wb_formula = load_workbook(filename=str(workbook_path), data_only=False)
    wb_values = load_workbook(filename=str(workbook_path), data_only=True)

    try:
        lines: List[str] = []

        for ws_formula in wb_formula.worksheets:
            sheet_name = ws_formula.title
            ws_values_sheet = wb_values[sheet_name]

            lines.append(f"[Sheet {sheet_name}]")

            for row in ws_formula.iter_rows():
                for cell_f in row:
                    addr = cell_f.coordinate
                    cell_v = ws_values_sheet[addr]

                    # Extract formula
                    formula = ""
                    if isinstance(cell_f.value, str) and cell_f.value.startswith("="):
                        formula = cell_f.value

                    # Extract value
                    value = cell_v.value

                    # Skip empty cells
                    if value is None and not formula:
                        continue

                    value_str = "" if value is None else str(value)

                    if formula:
                        lines.append(
                            f"{sheet_name} {addr}(value='{value_str}', formula='{formula}')"
                        )
                    else:
                        lines.append(
                            f"{sheet_name} {addr}(value='{value_str}')"
                        )

        text = "\n".join(lines)
        return text
    finally:
        try:
            wb_formula.close()
        except Exception:
            pass
        try:
            wb_values.close()
        except Exception:
            pass


# ==================== Rich Snapshot Builder ====================

def build_snapshot_rich(
    sheets_data: Dict[str, List[Dict[str, Any]]],
    changed_rows: Optional[Dict[str, Set[int]]] = None,
    changed_cols: Optional[Dict[str, Set[int]]] = None,
    max_rows_front: int = 10,
    max_rows_back: int = 10,
    max_chars: int = 500000,
) -> str:
    """
    Build a rich snapshot with selective row/column inclusion.
    
    Includes:
    - First max_rows_front rows
    - Last max_rows_back rows
    - All changed rows (if provided)
    - All rows containing changed columns (if provided)
    - All columns (no column limit)
    
    Args:
        sheets_data: Sheet data from read_workbook_data()
        changed_rows: Optional dict of changed row numbers per sheet
        changed_cols: Optional dict of changed column indices per sheet
        max_rows_front: Number of rows to keep from the start
        max_rows_back: Number of rows to keep from the end
        max_chars: Maximum characters for output
    
    Returns:
        Rich text snapshot
    """
    changed_rows = changed_rows or {}
    changed_cols = changed_cols or {}
    
    lines: List[str] = []
    sheet_names = sorted(sheets_data.keys())

    lines.append(
        "=== RICH SNAPSHOT: first/last rows + all changed rows/columns; all columns ==="
    )

    for sheet_name in sheet_names:
        items = sheets_data.get(sheet_name, []) or []
        
        # Build row map: row -> col_idx -> (addr, text, formula)
        row_map: Dict[int, Dict[int, Tuple[str, str, str]]] = {}

        for item in items:
            addr = item.get("address", "")
            if not addr:
                continue
            parsed = parse_addr(addr)
            if not parsed:
                continue
            row, col_idx = parsed

            text = str(item.get("text", "") or "")
            formula = str(item.get("formula", "") or "")
            row_map.setdefault(row, {})[col_idx] = (addr, text, formula)

        if not row_map:
            continue

        all_rows = sorted(row_map.keys())
        total_rows = len(all_rows)

        # Determine rows to keep
        changed_sheet_rows = changed_rows.get(sheet_name, set())
        changed_sheet_cols = changed_cols.get(sheet_name, set())

        # Start with first/last rows + changed rows
        base_rows = set(all_rows[:max_rows_front] + all_rows[-max_rows_back:])
        selected_rows = base_rows | (changed_sheet_rows & set(all_rows))

        # Add rows that contain changed columns
        for r in all_rows:
            cols = row_map.get(r, {})
            if any(c in changed_sheet_cols for c in cols.keys()):
                selected_rows.add(r)

        selected_rows_sorted = sorted(selected_rows)
        omitted_rows = total_rows - len(selected_rows_sorted)

        lines.append(f"[Sheet {sheet_name}] total rows with data: {total_rows}")
        if omitted_rows > 0:
            lines.append(f"... ({omitted_rows} rows omitted in this snapshot) ...")

        for r in selected_rows_sorted:
            cols = row_map.get(r, {})
            if not cols:
                continue
            cell_strs: List[str] = []
            for c in sorted(cols.keys()):
                addr, text, formula = cols[c]
                if formula:
                    cell_strs.append(f"{addr}(text='{text}', formula='{formula}')")
                else:
                    cell_strs.append(f"{addr}(text='{text}')")
            if cell_strs:
                lines.append(f"{sheet_name} Row {r}: " + "; ".join(cell_strs))

    text = "\n".join(lines)
    return text


# ==================== Simple Snapshot Builder ====================

def build_snapshot_simple(
    sheets_data: Dict[str, List[Dict[str, Any]]],
    changed_rows: Optional[Dict[str, Set[int]]] = None,
    changed_cols: Optional[Dict[str, Set[int]]] = None,
    max_rows_front: int = 10,
    max_rows_back: int = 10,
    max_cols: int = 5,
    max_chars: int = 500000,
) -> str:
    """
    Build a simple snapshot with limited columns (A-E by default).
    
    Includes:
    - First max_rows_front rows
    - Last max_rows_back rows
    - All changed rows (if provided)
    - First max_cols columns, plus any changed columns (if provided)
    
    Args:
        sheets_data: Sheet data from read_workbook_data()
        changed_rows: Optional dict of changed row numbers per sheet
        changed_cols: Optional dict of changed column indices per sheet
        max_rows_front: Number of rows to keep from the start
        max_rows_back: Number of rows to keep from the end
        max_cols: Maximum number of columns to include
        max_chars: Maximum characters for output
    
    Returns:
        Simple text snapshot
    """
    changed_rows = changed_rows or {}
    changed_cols = changed_cols or {}
    
    lines: List[str] = []
    sheet_names = sorted(sheets_data.keys())

    col_letter_end = index_to_col_letters(max_cols)
    lines.append(
        f"=== SIMPLE SNAPSHOT: first/last rows + changed rows; first {max_cols} columns A-{col_letter_end} (+ changed columns when provided) ==="
    )

    for sheet_name in sheet_names:
        items = sheets_data.get(sheet_name, []) or []
        
        # Build row map: row -> col_idx -> (addr, text, formula)
        row_map: Dict[int, Dict[int, Tuple[str, str, str]]] = {}
        changed_sheet_cols = changed_cols.get(sheet_name, set())
        keep_cols = set(range(1, max_cols + 1)) | set(changed_sheet_cols)

        for item in items:
            addr = item.get("address", "")
            if not addr:
                continue
            parsed = parse_addr(addr)
            if not parsed:
                continue
            row, col_idx = parsed

            # Skip columns not selected
            if col_idx not in keep_cols:
                continue

            text = str(item.get("text", "") or "")
            formula = str(item.get("formula", "") or "")
            row_map.setdefault(row, {})[col_idx] = (addr, text, formula)

        if not row_map:
            continue

        all_rows = sorted(row_map.keys())
        total_rows = len(all_rows)

        # Determine rows to keep
        changed_sheet_rows = changed_rows.get(sheet_name, set())
        base_rows = set(all_rows[:max_rows_front] + all_rows[-max_rows_back:])
        selected_rows = base_rows | (changed_sheet_rows & set(all_rows))

        selected_rows_sorted = sorted(selected_rows)
        omitted_rows = total_rows - len(selected_rows_sorted)

        lines.append(f"[Sheet {sheet_name}] total rows with data: {total_rows}")
        if omitted_rows > 0:
            lines.append(f"... ({omitted_rows} rows omitted in this snapshot) ...")

        for r in selected_rows_sorted:
            cols = row_map.get(r, {})
            if not cols:
                continue
            cell_strs: List[str] = []
            for c in sorted(cols.keys()):
                addr, text, formula = cols[c]
                if formula:
                    cell_strs.append(f"{addr}(text='{text}', formula='{formula}')")
                else:
                    cell_strs.append(f"{addr}(text='{text}')")
            if cell_strs:
                lines.append(f"{sheet_name} Row {r}: " + "; ".join(cell_strs))

    text = "\n".join(lines)
    return text


# ==================== JSON Export ====================

def build_snapshot_json(
    sheets_data: Dict[str, List[Dict[str, Any]]],
) -> Dict[str, Dict[str, Dict[str, Dict[str, str]]]]:
    """
    Build a JSON-structured snapshot of the workbook.
    
    Format:
        {
            "Sheet1": {
                "1": {
                    "A": {"text": "Hello", "formula": ""},
                    "B": {"text": "100", "formula": "=SUM(A1:A10)"}
                },
                "2": {
                    "A": {"text": "World"}
                }
            }
        }
    
    Args:
        sheets_data: Sheet data from read_workbook_data()
    
    Returns:
        Nested dictionary structure
    """
    result: Dict[str, Dict[str, Dict[str, Dict[str, str]]]] = {}

    for sheet_name, items in sheets_data.items():
        row_map: Dict[int, Dict[int, Tuple[str, str, str]]] = {}

        for item in items:
            addr = item.get("address", "")
            if not addr:
                continue
            parsed = parse_addr(addr)
            if not parsed:
                continue
            row, col_idx = parsed
            text = str(item.get("text", "") or "")
            formula = str(item.get("formula", "") or "")
            row_map.setdefault(row, {})[col_idx] = (addr, text, formula)

        if not row_map:
            continue

        sheet_obj: Dict[str, Dict[str, Dict[str, str]]] = {}
        for r in sorted(row_map.keys()):
            cols = row_map[r]
            if not cols:
                continue
            row_key = str(r)
            row_obj: Dict[str, Dict[str, str]] = {}
            for c, (_, text, formula) in sorted(cols.items()):
                col_letter = index_to_col_letters(c)
                cell_obj: Dict[str, str] = {}
                if text:
                    cell_obj["text"] = text
                if formula:
                    cell_obj["formula"] = formula
                if cell_obj:
                    row_obj[col_letter] = cell_obj
            if row_obj:
                sheet_obj[row_key] = row_obj

        if sheet_obj:
            result[sheet_name] = sheet_obj

    return result


# ==================== Compact Region Builder ====================

def build_region_text(
    sheet_items: List[Dict[str, Any]],
    max_chars: int = 8000,
) -> str:
    """
    Build compact text representation of sheet cells for analysis.
    Cells are sorted by row and column.
    
    Args:
        sheet_items: List of cell items for a single sheet
        max_chars: Maximum characters for output
    
    Returns:
        Compact text representation
    """
    rows: List[Tuple[int, int, str]] = []
    for item in sheet_items or []:
        addr = item.get("address", "")
        if not addr:
            continue
        parsed = parse_addr(addr)
        if not parsed:
            continue
        row, col_idx = parsed
        text = str(item.get("text", "") or "")
        formula = str(item.get("formula", "") or "")
        if formula:
            desc = f"{addr}(text='{text}', formula='{formula}')"
        else:
            desc = f"{addr}(text='{text}')"
        rows.append((row, col_idx, desc))

    rows.sort(key=lambda x: (x[0], x[1]))
    lines = [desc for _, _, desc in rows]
    text = "\n".join(lines)
    # No truncation here; let outer token_counter handle limits.
    return text


# ==================== Command Line Interface ====================

def main():
    parser = argparse.ArgumentParser(
        description="Build text-based snapshots of Excel workbooks"
    )
    parser.add_argument("input", type=Path, help="Path to the Excel file")
    parser.add_argument(
        "-m", "--mode", 
        choices=["full", "rich", "simple", "json"],
        default="full",
        help="Snapshot mode (default: full)"
    )
    parser.add_argument(
        "-o", "--output", 
        type=Path, 
        default=None,
        help="Output file path (default: print to console)"
    )
    parser.add_argument(
        "--max-chars", 
        type=int, 
        default=500000,
        help="Maximum characters for text output (default: 500000)"
    )
    parser.add_argument(
        "--max-rows-front", 
        type=int, 
        default=10,
        help="Number of rows to keep from start (for rich/simple modes, default: 10)"
    )
    parser.add_argument(
        "--max-rows-back", 
        type=int, 
        default=10,
        help="Number of rows to keep from end (for rich/simple modes, default: 10)"
    )
    parser.add_argument(
        "--max-cols", 
        type=int, 
        default=5,
        help="Maximum number of columns (for simple mode, default: 5)"
    )
    
    args = parser.parse_args()
    
    if not args.input.exists():
        print(f"[ERROR] File not found: {args.input}")
        return
    
    print(f"[INFO] Reading workbook: {args.input}")
    
    try:
        if args.mode == "full":
            # Full snapshot - directly read and build
            snapshot = build_snapshot_full(args.input, max_chars=args.max_chars)
            
        elif args.mode == "json":
            # JSON export
            sheets_data = read_workbook_data(args.input)
            snapshot_json = build_snapshot_json(sheets_data)
            snapshot = json.dumps(snapshot_json, ensure_ascii=False, indent=2)
            
        elif args.mode == "rich":
            # Rich snapshot
            sheets_data = read_workbook_data(args.input)
            snapshot = build_snapshot_rich(
                sheets_data,
                max_rows_front=args.max_rows_front,
                max_rows_back=args.max_rows_back,
                max_chars=args.max_chars
            )
            
        elif args.mode == "simple":
            # Simple snapshot
            sheets_data = read_workbook_data(args.input)
            snapshot = build_snapshot_simple(
                sheets_data,
                max_rows_front=args.max_rows_front,
                max_rows_back=args.max_rows_back,
                max_cols=args.max_cols,
                max_chars=args.max_chars
            )
        
        else:
            print(f"[ERROR] Unknown mode: {args.mode}")
            return
        
        # Output result
        if args.output:
            args.output.parent.mkdir(parents=True, exist_ok=True)
            with args.output.open("w", encoding="utf-8") as f:
                f.write(snapshot)
            print(f"[SUCCESS] Snapshot saved to: {args.output}")
        else:
            print("\n" + "="*60)
            print(snapshot)
            print("="*60)
        
        # Print statistics
        print(f"\n[INFO] Snapshot size: {len(snapshot)} characters")
        
    except Exception as e:
        print(f"[ERROR] Failed to build snapshot: {e}")
        import traceback
        traceback.print_exc()


if __name__ == "__main__":
    main()
