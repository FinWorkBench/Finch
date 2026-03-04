"""
Excel Content Builder

Builds content parts for Excel evaluation tasks.
"""

import hashlib
import json
import logging
import re
from collections import defaultdict
from pathlib import Path
from typing import List, Dict, Optional, Set, Any, Tuple

from .config import (
    Captions,
    MAX_TEXT_CHARS,
    EXCEL_SNAPSHOT_MAX_CHARS,
    EXCEL_SNAPSHOT_MAX_ROWS_FRONT,
    EXCEL_SNAPSHOT_MAX_ROWS_BACK,
    EXCEL_SNAPSHOT_MAX_COLS_SIMPLE,
)

from .cache_manager import CacheManager
from .token_counter import (
    can_add_text_within_char_limit,
)

from .utils import (
    image_to_data_url,
    extract_changed_sheets,
    format_diff_summary,
    get_workbook_sheetnames_fast
)

from dataclasses import dataclass, field


logger = logging.getLogger(__name__)

@dataclass
class ExcelSideResult:
    """Side-level info used to decide source focus sheets and avoid duplicated source additions."""
    exists: bool = False
    used_diff: bool = False
    changed_sheets: Set[str] = field(default_factory=set)
    changed_rows: Dict[str, Set[int]] = field(default_factory=dict)
    changed_cols: Dict[str, Set[int]] = field(default_factory=dict)

    @staticmethod
    def empty() -> "ExcelSideResult":
        return ExcelSideResult(False, False, set(), {}, {})

class ExcelContentBuilder:
    """
    Builds content parts for Excel evaluation tasks.
    
    Handles:
    - Diff computation and caching
    - Snapshot generation and caching
    - Screenshot generation and caching
    - Intelligent selection between full snapshot vs diff
    - text char limit management
    """
    
    def __init__(
        self,
        task_dir: Path,
        cache_manager: CacheManager,
        excel_comparator_path: str = "excel_comparator.py",
        excel_snapshot_builder_path: str = "excel_snapshot_builder.py",
        excel_shot_generator_path: str = "excel_shot_generator.py",
    ):
        """
        Initialize Excel content builder.
        
        Args:
            task_dir: Path to task directory
            cache_manager: CacheManager instance
            excel_comparator_path: Path to excel_comparator module
            excel_snapshot_builder_path: Path to excel_snapshot_builder module
            excel_shot_generator_path: Path to excel_shot_generator module
        """
        self.task_dir = Path(task_dir)
        self.cache = cache_manager
        self.comparator_path = excel_comparator_path
        self.snapshot_builder_path = excel_snapshot_builder_path
        self.shot_generator_path = excel_shot_generator_path
    
    def build_side(
        self,
        content_parts: List[Dict[str, Any]],
        side: str,                       # "reference" | "output"
        target_file: Optional[Path],
        source_file: Optional[Path] = None,
    ) -> ExcelSideResult:
        """
        Build one side (reference or output) independently.

        - If source_file exists: compute diff(source -> target), then pick diff summary or full snapshot
          based on shorter text, and attach screenshots (selected if diff, else all).
        - If no source_file: attach full snapshot + all screenshots.
        """
        side = (side or "").lower().strip()
        assert side in {"reference", "output"}

        if not target_file or not Path(target_file).exists():
            # Keep going even if one side is missing
            if side == "reference":
                content_parts.append({
                    "type": "text",
                    "text": Captions.EXCEL_REFERENCE_MISSING.format(path=str(target_file))
                })
            else:
                content_parts.append({
                    "type": "text",
                    "text": Captions.EXCEL_OUTPUT_MISSING.format(path=str(target_file))
                })
            return ExcelSideResult.empty()

        target_file = Path(target_file)
        result = ExcelSideResult(exists=True)

        # No source: always full snapshot + all screenshots
        if not source_file or not Path(source_file).exists():
            full_snapshot = self._get_full_snapshot(target_file)

            if side == "reference":
                content_parts.append({"type": "text", "text": Captions.EXCEL_REFERENCE_FULL.format(snapshot=full_snapshot)})
                shots = self._get_all_screenshots(target_file)
                for s in shots:
                    content_parts.append({"type": "text", "text": Captions.EXCEL_REFERENCE_SHEET.format(sheet_name=s["sheet_name"])})
                    content_parts.append({"type": "image_url", "image_url": {"url": image_to_data_url(Path(s["path"]))}})
            else:
                content_parts.append({"type": "text", "text": Captions.EXCEL_OUTPUT_FULL.format(snapshot=full_snapshot)})
                shots = self._get_all_screenshots(target_file)
                for s in shots:
                    content_parts.append({"type": "text", "text": Captions.EXCEL_OUTPUT_SHEET.format(sheet_name=s["sheet_name"])})
                    content_parts.append({"type": "image_url", "image_url": {"url": image_to_data_url(Path(s["path"]))}})

            return result

        # With source: diff-based decision (diff summary vs full snapshot)
        source_file = Path(source_file)
        diff = self._get_diff(source_file, target_file)
        changed = set(extract_changed_sheets(diff, target_file))
        result.changed_sheets = changed
        changed_rows, changed_cols = self._extract_changed_rows_cols(diff)
        result.changed_rows = changed_rows
        result.changed_cols = changed_cols

        full_snapshot = self._get_full_snapshot(target_file)
        diff_text = format_diff_summary(diff)
        use_diff = len(diff_text) < len(full_snapshot)
        result.used_diff = use_diff

        if side == "reference":
            if use_diff:
                content_parts.append({"type": "text", "text": Captions.EXCEL_REFERENCE_DIFF.format(diff=diff_text)})
                shots = self._get_selected_screenshots(target_file, sorted(changed))
            else:
                content_parts.append({"type": "text", "text": Captions.EXCEL_REFERENCE_FULL.format(snapshot=full_snapshot)})
                shots = self._get_all_screenshots(target_file)

            for s in shots:
                content_parts.append({"type": "text", "text": Captions.EXCEL_REFERENCE_SHEET.format(sheet_name=s["sheet_name"])})
                content_parts.append({"type": "image_url", "image_url": {"url": image_to_data_url(Path(s["path"]))}})

        else:
            if use_diff:
                content_parts.append({"type": "text", "text": Captions.EXCEL_OUTPUT_DIFF.format(diff=diff_text)})
                shots = self._get_selected_screenshots(target_file, sorted(changed))
            else:
                content_parts.append({"type": "text", "text": Captions.EXCEL_OUTPUT_FULL.format(snapshot=full_snapshot)})
                shots = self._get_all_screenshots(target_file)

            for s in shots:
                content_parts.append({"type": "text", "text": Captions.EXCEL_OUTPUT_SHEET.format(sheet_name=s["sheet_name"])})
                content_parts.append({"type": "image_url", "image_url": {"url": image_to_data_url(Path(s["path"]))}})

        return result

    def add_source_once(
        self,
        content_parts: List[Dict[str, Any]],
        source_file: Path,
        focus_sheets: Set[str],
        changed_rows: Optional[Dict[str, Set[int]]] = None,
        changed_cols: Optional[Dict[str, Set[int]]] = None,
    ) -> None:
        """
        Append source snapshot + screenshots exactly once.

        If focus_sheets is non-empty, prefer selected screenshots for those sheets (filtered to sheets existing in source).
        Otherwise, attach screenshots for all sheets.
        """
        source_file = Path(source_file)
        if not source_file.exists():
            return

        # Snapshot: prefer rich snapshot unless text char limit would be exceeded
        source_rich_snapshot = self._get_rich_snapshot(
            source_file,
            changed_rows=changed_rows,
            changed_cols=changed_cols,
        )
        rich_text = Captions.EXCEL_SOURCE_RICH.format(snapshot=source_rich_snapshot)

        if can_add_text_within_char_limit(content_parts, rich_text, MAX_TEXT_CHARS):
            content_parts.append({"type": "text", "text": rich_text})
        else:
            source_simple_snapshot = self._get_simple_snapshot(
                source_file,
                changed_rows=changed_rows,
                changed_cols=changed_cols,
            )
            simple_text = Captions.EXCEL_SOURCE_SIMPLE.format(snapshot=source_simple_snapshot)
            content_parts.append({"type": "text", "text": simple_text})
            logger.info("Using simple snapshot for source due to text char limit")

        # Screenshots: selected vs all
        try:
            use_focus = set(focus_sheets or set())

            if use_focus:
                try:
                    names = set(get_workbook_sheetnames_fast(source_file))
                    use_focus = {s for s in use_focus if s in names}
                except Exception as e:
                    logger.warning(f"Failed to read source sheetnames fast; will try screenshot anyway: {e}")

            shots = (
                self._get_selected_screenshots(source_file, sorted(use_focus))
                if use_focus
                else self._get_all_screenshots(source_file)
            )

            for s in shots:
                caption = (
                    Captions.EXCEL_SOURCE_SHEET.format(sheet_name=s["sheet_name"])
                    if hasattr(Captions, "EXCEL_SOURCE_SHEET")
                    else f"Source sheet: {s['sheet_name']}"
                )
                content_parts.append({"type": "text", "text": caption})
                content_parts.append({"type": "image_url", "image_url": {"url": image_to_data_url(Path(s["path"]))}})

        except Exception as e:
            logger.warning(f"Failed to add source screenshots: {e}")
        



    
    # ==================== Helper Methods ====================

    @staticmethod
    def _parse_addr(addr: str) -> Optional[Tuple[int, int]]:
        """
        Parse A1-style address into (row, col_idx), where col_idx is 1-based.
        Returns None if parsing fails.
        """
        if not addr:
            return None
        m = re.match(r"^([A-Za-z]+)(\d+)$", addr)
        if not m:
            return None
        col_letters, row_str = m.groups()
        row = int(row_str)
        col_idx = 0
        for ch in col_letters.upper():
            col_idx = col_idx * 26 + (ord(ch) - ord("A") + 1)
        return row, col_idx

    @classmethod
    def _extract_changed_rows_cols(
        cls, diff: Dict[str, Any]
    ) -> Tuple[Dict[str, Set[int]], Dict[str, Set[int]]]:
        """
        Extract changed row/col indices per sheet from diff result.
        Only uses cell-level diffs (text/formula).
        """
        rows: Dict[str, Set[int]] = defaultdict(set)
        cols: Dict[str, Set[int]] = defaultdict(set)

        sheets_diff = diff.get("sheets_diff", {}) or {}
        for sheet_name, sheet_info in sheets_diff.items():
            for change in (sheet_info.get("diff") or []):
                addr = change.get("address", "")
                parsed = cls._parse_addr(addr)
                if not parsed:
                    continue
                r, c = parsed
                rows[sheet_name].add(r)
                cols[sheet_name].add(c)

        return dict(rows), dict(cols)

    @staticmethod
    def _snapshot_cache_key(
        changed_rows: Optional[Dict[str, Set[int]]],
        changed_cols: Optional[Dict[str, Set[int]]],
    ) -> Optional[str]:
        if not changed_rows and not changed_cols:
            return None

        def _normalize(d: Optional[Dict[str, Set[int]]]) -> Dict[str, List[int]]:
            out: Dict[str, List[int]] = {}
            for k, v in (d or {}).items():
                if v:
                    out[k] = sorted(set(v))
            return out

        payload = {
            "rows": _normalize(changed_rows),
            "cols": _normalize(changed_cols),
        }
        blob = json.dumps(payload, sort_keys=True, separators=(",", ":"))
        digest = hashlib.sha1(blob.encode("utf-8")).hexdigest()[:10]
        return f"rc_{digest}"
    
    def _get_diff(self, source_file: Path, target_file: Path) -> dict:
        """Get or compute diff between source and target files."""
        # Check cache first
        cached_diff = self.cache.get_diff_cache(source_file.name, target_file.name)
        if cached_diff:
            return cached_diff
        
        # Compute diff using excel_comparator
        logger.info(f"Computing diff: {source_file.name} -> {target_file.name}")
        
        import sys
        sys.path.insert(0, str(Path(self.comparator_path).parent))
        from ..excel_extract.excel_comparator import get_diff_info
        
        diff = get_diff_info(source_file, target_file)
        
        # Cache result
        self.cache.save_diff_cache(source_file.name, target_file.name, diff)
        
        return diff
    
    def _get_full_snapshot(self, excel_file: Path) -> str:
        """Get or compute full snapshot of Excel file."""
        # Check cache
        cached = self.cache.get_snapshot_cache(excel_file.name, "full")
        if cached:
            return cached
        
        # Generate snapshot
        logger.info(f"Generating full snapshot: {excel_file.name}")
        
        import sys
        sys.path.insert(0, str(Path(self.snapshot_builder_path).parent))
        from ..excel_extract.excel_snapshot_builder import build_snapshot_full
        
        snapshot = build_snapshot_full(excel_file, max_chars=EXCEL_SNAPSHOT_MAX_CHARS)
        
        # Cache result
        self.cache.save_snapshot_cache(excel_file.name, "full", snapshot)
        
        return snapshot
    
    def _get_rich_snapshot(
        self,
        excel_file: Path,
        changed_rows: Optional[Dict[str, Set[int]]] = None,
        changed_cols: Optional[Dict[str, Set[int]]] = None,
    ) -> str:
        """Get or compute rich snapshot of Excel file."""
        cache_key = self._snapshot_cache_key(changed_rows, changed_cols)
        cache_name = excel_file.name if not cache_key else f"{excel_file.name}__{cache_key}"
        cached = self.cache.get_snapshot_cache(cache_name, "rich")
        if cached:
            return cached
        
        # Generate snapshot
        logger.info(f"Generating rich snapshot: {excel_file.name}")
        
        import sys
        sys.path.insert(0, str(Path(self.snapshot_builder_path).parent))
        from ..excel_extract.excel_snapshot_builder import read_workbook_data, build_snapshot_rich
        
        data = read_workbook_data(excel_file)
        snapshot = build_snapshot_rich(
            data,
            changed_rows=changed_rows,
            changed_cols=changed_cols,
            max_rows_front=EXCEL_SNAPSHOT_MAX_ROWS_FRONT,
            max_rows_back=EXCEL_SNAPSHOT_MAX_ROWS_BACK,
            max_chars=EXCEL_SNAPSHOT_MAX_CHARS
        )

        # Cache result
        self.cache.save_snapshot_cache(cache_name, "rich", snapshot)
        
        return snapshot
    
    def _get_simple_snapshot(
        self,
        excel_file: Path,
        changed_rows: Optional[Dict[str, Set[int]]] = None,
        changed_cols: Optional[Dict[str, Set[int]]] = None,
    ) -> str:
        """Get or compute simple snapshot of Excel file."""
        cache_key = self._snapshot_cache_key(changed_rows, changed_cols)
        cache_name = excel_file.name if not cache_key else f"{excel_file.name}__{cache_key}"
        cached = self.cache.get_snapshot_cache(cache_name, "simple")
        if cached:
            return cached
        
        # Generate snapshot
        logger.info(f"Generating simple snapshot: {excel_file.name}")
        
        import sys
        sys.path.insert(0, str(Path(self.snapshot_builder_path).parent))
        from ..excel_extract.excel_snapshot_builder import read_workbook_data, build_snapshot_simple
        
        data = read_workbook_data(excel_file)
        snapshot = build_snapshot_simple(
            data,
            changed_rows=changed_rows,
            changed_cols=changed_cols,
            max_rows_front=EXCEL_SNAPSHOT_MAX_ROWS_FRONT,
            max_rows_back=EXCEL_SNAPSHOT_MAX_ROWS_BACK,
            max_cols=EXCEL_SNAPSHOT_MAX_COLS_SIMPLE,
            max_chars=EXCEL_SNAPSHOT_MAX_CHARS
        )

        # Cache result
        self.cache.save_snapshot_cache(cache_name, "simple", snapshot)
        
        return snapshot
    
    def _get_all_screenshots(self, excel_file: Path) -> List[Dict[str, str]]:
        """Get or generate screenshots for all sheets."""
        # First, get all sheet names
        import sys
        sys.path.insert(0, str(Path(self.comparator_path).parent))
        import openpyxl
        wb = openpyxl.load_workbook(excel_file, data_only=True)
        sheet_names = [ws.title for ws in wb.worksheets]
        wb.close()
        
        # Check cache
        cached = self.cache.get_screenshots_cache(excel_file.name, sheet_names)
        if cached:
            return cached
        
        # Generate screenshots
        logger.info(f"Generating screenshots for all sheets: {excel_file.name}")
        
        sys.path.insert(0, str(Path(self.shot_generator_path).parent))
        from ..excel_extract.excel_shot_generator import ExcelShotGenerator
        
        generator = ExcelShotGenerator()
        screenshots = generator.generate_screenshots(str(excel_file), sheet_names=None)
        
        # Cache results
        self.cache.save_screenshots_cache(excel_file.name, screenshots)
        
        return screenshots
    
    def _get_selected_screenshots(
        self,
        excel_file: Path,
        sheet_names: List[str]
    ) -> List[Dict[str, str]]:
        """Get or generate screenshots for selected sheets."""
        if not sheet_names:
            return []
        
        # Check cache
        cached = self.cache.get_screenshots_cache(excel_file.name, sheet_names)
        if cached:
            return cached
        
        # Generate screenshots
        logger.info(f"Generating screenshots for {len(sheet_names)} sheets: {excel_file.name}")
        
        import sys
        sys.path.insert(0, str(Path(self.shot_generator_path).parent))
        from ..excel_extract.excel_shot_generator import ExcelShotGenerator
        
        generator = ExcelShotGenerator()
        screenshots = generator.generate_screenshots(str(excel_file), sheet_names=sheet_names)
        
        # Cache results
        self.cache.save_screenshots_cache(excel_file.name, screenshots)
        
        return screenshots
