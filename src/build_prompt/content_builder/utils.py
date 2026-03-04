"""
Utility Functions for Content Builder
"""

import base64
import logging
from pathlib import Path
from typing import List, Optional, Set

from .config import EXCEL_EXTENSIONS, IMAGE_EXTENSIONS
import openpyxl
from openpyxl.chartsheet import Chartsheet

logger = logging.getLogger(__name__)


def get_workbook_sheetnames_fast(xlsx_path: Path) -> Set[str]:
    """Fast sheetname listing without launching Excel."""
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=False)
    names = set()
    for ws in getattr(wb, "_sheets", []):
        # Optional: skip chart sheets if your screenshot tool can’t handle them
        if isinstance(ws, Chartsheet):
            continue
        names.add(ws.title)
    try:
        wb.close()
    except Exception:
        pass
    return names

def has_excel_files(file_list: List[str]) -> bool:
    """
    Check if file list contains any Excel files.
    
    Args:
        file_list: List of file paths
    
    Returns:
        True if any Excel files found
    """
    return any(
        Path(f).suffix.lower() in EXCEL_EXTENSIONS
        for f in file_list
    )


def get_excel_files(file_list: List[str]) -> List[Path]:
    """
    Filter and return Excel files from file list.
    
    Args:
        file_list: List of file paths
    
    Returns:
        List of Path objects for Excel files
    """
    return [
        Path(f) for f in file_list
        if Path(f).suffix.lower() in EXCEL_EXTENSIONS
    ]


def get_first_excel_file(file_list: List[str]) -> Optional[Path]:
    """
    Get the first Excel file from file list.
    
    Args:
        file_list: List of file paths
    
    Returns:
        Path to first Excel file or None
    """
    excel_files = get_excel_files(file_list)
    return excel_files[0] if excel_files else None


def image_to_data_url(image_path: Path) -> str:
    """
    Convert image file to base64 data URL.
    
    Args:
        image_path: Path to image file
    
    Returns:
        Data URL string (e.g., "data:image/png;base64,...")
    """
    with open(image_path, "rb") as f:
        image_data = f.read()
    
    base64_data = base64.b64encode(image_data).decode("utf-8")
    
    # Determine MIME type from extension
    ext = image_path.suffix.lower()
    mime_type_map = {
        ".png": "image/png",
        ".jpg": "image/jpeg",
        ".jpeg": "image/jpeg",
        ".gif": "image/gif",
        ".bmp": "image/bmp",
    }
    mime_type = mime_type_map.get(ext, "image/png")
    
    return f"data:{mime_type};base64,{base64_data}"




def extract_changed_sheets(diff: dict, target_file: Optional[Path] = None) -> Set[str]:
    """
    Extract changed sheet names from diff result.
    If target_file is provided, only return sheets that exist in target_file.
    """
    summary = diff.get("summary", {}) or {}
    all_changed = summary.get("all_changed", []) or []

    # all_changed is expected to be a list[str]
    changed = set(all_changed)

    if target_file is not None:
        target_sheets = get_workbook_sheetnames_fast(Path(target_file))
        changed = changed & target_sheets

    return changed



def format_diff_summary(diff: dict) -> str:
    """
    Format diff result as a readable text summary.
    
    Args:
        diff: Diff dict from excel_comparator
    
    Returns:
        Formatted text summary
    """
    lines = []
    lines.append("=== Changes Summary ===\n")
    
    sheets_diff = diff.get("sheets_diff", {})
    
    for sheet_name in sorted(sheets_diff.keys()):
        sheet_info = sheets_diff[sheet_name]
        diffs = sheet_info.get("diff", [])
        
        if not diffs:
            continue
        
        lines.append(f"\n[Sheet: {sheet_name}] ({len(diffs)} changes)")
        
        for change in diffs:
            addr = change.get("address", "")
            tb = change.get("text_before", "")
            ta = change.get("text_after", "")
            fb = change.get("formula_before", "")
            fa = change.get("formula_after", "")
            
            parts = []
            if tb != ta:
                parts.append(f"text: '{tb}' → '{ta}'")
            if fb != fa:
                parts.append(f"formula: '{fb}' → '{fa}'")
            
            lines.append(f"  {addr}: {'; '.join(parts)}")
        
    
    if not any(sheets_diff.get(s, {}).get("diff") for s in sheets_diff):
        lines.append("\nNo changes detected.")
    
    return "\n".join(lines)


def sanitize_filename(name: str) -> str:
    """
    Sanitize string for use as filename.
    
    Args:
        name: Original name
    
    Returns:
        Sanitized name safe for filesystem
    """
    return "".join(
        c if c.isalnum() or c in (' ', '-', '_') else '_'
        for c in name
    )


def ensure_directory(path: Path) -> Path:
    """
    Ensure directory exists, create if necessary.
    
    Args:
        path: Directory path
    
    Returns:
        Path object (guaranteed to exist)
    """
    path = Path(path)
    path.mkdir(parents=True, exist_ok=True)
    return path
