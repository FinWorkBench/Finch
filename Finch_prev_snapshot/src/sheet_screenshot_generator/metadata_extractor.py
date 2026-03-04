import os
import json
from pathlib import Path
from typing import List, Dict, Any

import openpyxl
from openpyxl.chart import BarChart, LineChart, PieChart, ScatterChart, AreaChart
from openpyxl.chartsheet import Chartsheet


class EnhancedExcelExtractor:
    """Enhanced extractor with comprehensive Excel content analysis."""

    def __init__(self, xlsx_path: str):
        self.xlsx_path = xlsx_path
        try:
            self.workbook = openpyxl.load_workbook(xlsx_path, data_only=False)
        except Exception as e:
            error_msg = str(e)
            # Detect old Excel 97-2003 format pretending to be .xlsx
            if "not a zip file" in error_msg.lower():
                raise ValueError(
                    f"Cannot open Excel file '{os.path.basename(xlsx_path)}': "
                    f"File appears to be in old Excel 97-2003 format (.xls) but has .xlsx extension. "
                    f"openpyxl cannot read this format. Please convert to .xlsx or rename to .xls"
                )
            else:
                raise ValueError(
                    f"Cannot open Excel file '{os.path.basename(xlsx_path)}': {error_msg}. "
                    f"File may be corrupted or in use."
                )

    def extract_all_metadata(self) -> Dict[str, Any]:
        """Extract comprehensive metadata from all sheets in the workbook."""
        metadata: Dict[str, Any] = {
            "file_name": Path(self.xlsx_path).name,
            "file_path": str(self.xlsx_path),
            "sheet_count": len(self.workbook.sheetnames),
            "sheet_names": self.workbook.sheetnames,
            "sheets": [],
        }

        for sheet_name in self.workbook.sheetnames:
            sheet_data = self.extract_sheet_content(sheet_name)
            metadata["sheets"].append(sheet_data)

        return metadata

    def extract_sheet_content(self, sheet_name: str | None = None) -> Dict[str, Any]:
        """Extract comprehensive content from a sheet."""
        ws = self.workbook[sheet_name] if sheet_name else self.workbook.active

        # Chartsheet has no cell grid; handle separately
        if isinstance(ws, Chartsheet):
            print(
                f"   !! EnhancedExcelExtractor: sheet '{ws.title}' in "
                f"'{os.path.basename(self.xlsx_path)}' is a chart sheet; "
                f"skip cell-based metadata and only mark it as chart_sheet."
            )
            return {
                "sheet_name": ws.title,
                "sheet_type": "chart_sheet",
                "dimensions": {"rows": None, "columns": None},
                "data": [],
                "headers": {"detected": False},
                "formatting": {},
                "formulas": {},
                "charts": [],
                "conditional_formatting": {},
                "data_summary": {"empty": True},
                "number_formats": {},
                "table_structure": {
                    "has_multiple_sheets": len(self.workbook.sheetnames) > 1,
                    "sheet_count": len(self.workbook.sheetnames),
                    "appears_tabular": False,
                    "density": 0.0,
                },
            }

        content: Dict[str, Any] = {
            "sheet_name": ws.title,
            "dimensions": {"rows": ws.max_row, "columns": ws.max_column},
            "data": self._extract_data_with_types(ws),
            "headers": self._detect_headers(ws),
            "formatting": self._extract_comprehensive_formatting(ws),
            "formulas": self._extract_formulas_detailed(ws),
            "charts": self._extract_charts_detailed(ws),
            "conditional_formatting": self._extract_conditional_formatting_detailed(ws),
            "data_summary": self._summarize_data(ws),
            "number_formats": self._extract_number_formats(ws),
            "table_structure": self._analyze_table_structure(ws),
        }

        return content

    # ---------- Cell data & headers ----------

    def _extract_data_with_types(self, ws) -> List[List[Dict[str, Any]]]:
        """Extract data with cell types and number formats (limited rows/cols)."""
        data: List[List[Dict[str, Any]]] = []
        max_rows = min(ws.max_row, 100)
        max_cols = min(ws.max_column, 25)

        for row in ws.iter_rows(min_row=1, max_row=max_rows, max_col=max_cols):
            row_data = []
            for cell in row:
                cell_info: Dict[str, Any] = {
                    "value": cell.value,
                    "address": cell.coordinate,
                    "data_type": cell.data_type,
                    "number_format": cell.number_format or "General",
                }

                if cell.data_type == "f":
                    cell_info["formula"] = cell.value
                    try:
                        # For now we just duplicate value as calculated_value
                        cell_info["calculated_value"] = cell.value
                    except Exception:
                        pass

                row_data.append(cell_info)
            data.append(row_data)

        return data

    def _detect_headers(self, ws) -> Dict[str, Any]:
        """Detect potential header row in the first row."""
        if ws.max_row == 0:
            return {"detected": False}

        first_row_cells = list(
            ws.iter_rows(min_row=1, max_row=1, max_col=min(ws.max_column, 20))
        )[0]

        has_bold = any(cell.font and cell.font.bold for cell in first_row_cells if cell.value)
        has_values = any(cell.value for cell in first_row_cells)
        header_values = [cell.value for cell in first_row_cells if cell.value]

        return {
            "detected": has_bold or (has_values and ws.max_row > 1),
            "likely_header_row": 1 if (has_bold or has_values) else None,
            "header_values": header_values[:15],
            "has_bold_formatting": has_bold,
        }

    # ---------- Formatting & styles ----------

    def _extract_comprehensive_formatting(self, ws) -> Dict[str, Any]:
        """Extract detailed formatting information from the first N rows/cols."""
        formatting: Dict[str, Any] = {
            "bold_cells": [],
            "italic_cells": [],
            "underline_cells": [],
            "colored_cells": [],
            "highlighted_cells": [],
            "merged_cells": [],
            "aligned_cells": {},
            "bordered_cells": [],
            "font_sizes": {},
            "color_summary": {},
        }

        max_rows = min(ws.max_row, 50)
        max_cols = min(ws.max_column, 20)

        colors_used: Dict[str, int] = {}
        bg_colors_used: Dict[str, int] = {}

        for row in ws.iter_rows(min_row=1, max_row=max_rows, max_col=max_cols):
            for cell in row:
                if not cell.value:
                    continue

                if cell.font:
                    if cell.font.bold:
                        formatting["bold_cells"].append(cell.coordinate)
                    if cell.font.italic:
                        formatting["italic_cells"].append(cell.coordinate)
                    if cell.font.underline:
                        formatting["underline_cells"].append(cell.coordinate)

                    if cell.font.size:
                        size = cell.font.size
                        formatting["font_sizes"].setdefault(size, []).append(cell.coordinate)

                    if cell.font.color and str(cell.font.color.rgb) != "FF000000":
                        color = str(cell.font.color.rgb)
                        colors_used[color] = colors_used.get(color, 0) + 1
                        formatting["colored_cells"].append(
                            {"cell": cell.coordinate, "color": color}
                        )

                if cell.fill and getattr(cell.fill, "patternType", None) not in (
                    None,
                    "none",
                ):
                    color = (
                        str(cell.fill.fgColor.rgb)
                        if getattr(cell.fill, "fgColor", None)
                        else "unknown"
                    )
                    bg_colors_used[color] = bg_colors_used.get(color, 0) + 1
                    formatting["highlighted_cells"].append(
                        {
                            "cell": cell.coordinate,
                            "color": color,
                            "pattern": cell.fill.patternType,
                        }
                    )

                if cell.alignment:
                    align_key = f"{cell.alignment.horizontal}_{cell.alignment.vertical}"
                    if align_key != "None_None":
                        formatting["aligned_cells"].setdefault(align_key, []).append(
                            cell.coordinate
                        )

                if cell.border and any(
                    [cell.border.left, cell.border.right, cell.border.top, cell.border.bottom]
                ):
                    formatting["bordered_cells"].append(cell.coordinate)

        formatting["merged_cells"] = [str(cr) for cr in ws.merged_cells.ranges]

        formatting["color_summary"] = {
            "font_colors_used": len(colors_used),
            "background_colors_used": len(bg_colors_used),
            "most_common_font_color": max(colors_used.items(), key=lambda x: x[1])[0]
            if colors_used
            else None,
            "most_common_bg_color": max(bg_colors_used.items(), key=lambda x: x[1])[0]
            if bg_colors_used
            else None,
        }

        return formatting

    # ---------- Formulas & charts & CF ----------

    def _extract_formulas_detailed(self, ws) -> Dict[str, Any]:
        """Extract formulas and their rough types."""
        formulas: List[Dict[str, Any]] = []
        formula_types: Dict[str, int] = {}

        max_rows = min(ws.max_row, 100)
        max_cols = min(ws.max_column, 20)

        for row in ws.iter_rows(min_row=1, max_row=max_rows, max_col=max_cols):
            for cell in row:
                if cell.data_type == "f" and cell.value:
                    formula_str = str(cell.value)
                    formula_type = self._categorize_formula(formula_str)
                    formula_types[formula_type] = formula_types.get(formula_type, 0) + 1

                    formulas.append(
                        {
                            "cell": cell.coordinate,
                            "formula": formula_str,
                            "type": formula_type,
                        }
                    )

        return {
            "formulas": formulas[:50],
            "count": len(formulas),
            "formula_types": formula_types,
        }

    def _categorize_formula(self, formula: str) -> str:
        """Categorize formula by simple heuristics."""
        formula_upper = formula.upper()

        if "SUM(" in formula_upper:
            return "sum"
        if "AVERAGE(" in formula_upper or "AVG(" in formula_upper:
            return "average"
        if "COUNT(" in formula_upper:
            return "count"
        if "IF(" in formula_upper:
            return "conditional"
        if any(k in formula_upper for k in ("VLOOKUP(", "HLOOKUP(", "XLOOKUP(")):
            return "lookup"
        if "%" in formula or "*100" in formula:
            return "percentage"
        if any(op in formula for op in ["+", "-", "*", "/"]):
            return "arithmetic"
        return "other"

    def _extract_charts_detailed(self, ws) -> List[Dict[str, Any]]:
        """Extract detailed chart information."""
        charts: List[Dict[str, Any]] = []

        for chart in getattr(ws, "_charts", []):
            chart_info: Dict[str, Any] = {
                "type": type(chart).__name__,
                "title": None,
                "data_range": None,
                "series_count": 0,
                "position": None,
            }

            if getattr(chart, "title", None):
                chart_info["title"] = str(chart.title)

            if getattr(chart, "series", None):
                chart_info["series_count"] = len(chart.series)

            if getattr(chart, "anchor", None):
                chart_info["position"] = str(chart.anchor)

            if isinstance(chart, BarChart):
                chart_info["chart_subtype"] = "bar"
                if hasattr(chart, "type"):
                    chart_info["bar_type"] = chart.type
            elif isinstance(chart, LineChart):
                chart_info["chart_subtype"] = "line"
            elif isinstance(chart, PieChart):
                chart_info["chart_subtype"] = "pie"
            elif isinstance(chart, ScatterChart):
                chart_info["chart_subtype"] = "scatter"
            elif isinstance(chart, AreaChart):
                chart_info["chart_subtype"] = "area"

            charts.append(chart_info)

        return charts

    def _extract_conditional_formatting_detailed(self, ws) -> Dict[str, Any]:
        """Extract conditional formatting rules."""
        cond_formats: List[Dict[str, Any]] = []

        for range_string, rules in ws.conditional_formatting._cf_rules.items():
            for rule in rules:
                rule_info: Dict[str, Any] = {
                    "range": range_string,
                    "type": type(rule).__name__,
                    "rule_details": {},
                }

                if hasattr(rule, "type"):
                    rule_info["rule_details"]["cf_type"] = rule.type
                if hasattr(rule, "operator"):
                    rule_info["rule_details"]["operator"] = rule.operator
                if hasattr(rule, "formula"):
                    rule_info["rule_details"]["formula"] = str(rule.formula)

                if "ColorScale" in type(rule).__name__:
                    rule_info["rule_details"]["scale_type"] = "color_scale"
                if "DataBar" in type(rule).__name__:
                    rule_info["rule_details"]["scale_type"] = "data_bar"
                if "IconSet" in type(rule).__name__:
                    rule_info["rule_details"]["scale_type"] = "icon_set"

                cond_formats.append(rule_info)

        return {
            "rules": cond_formats,
            "count": len(cond_formats),
            "ranges_affected": len({cf["range"] for cf in cond_formats}),
        }

    # ---------- Data summary & structure ----------

    def _summarize_data(self, ws) -> Dict[str, Any]:
        """Summarize data characteristics."""
        if ws.max_row == 0:
            return {"empty": True}

        numeric_cols = []
        text_cols = []
        date_cols = []  # placeholder, you can extend it if needed

        max_check_rows = min(ws.max_row, 20)
        max_cols = min(ws.max_column, 20)

        for col_idx in range(1, max_cols + 1):
            col_values = []
            for row_idx in range(2, max_check_rows + 1):
                cell = ws.cell(row_idx, col_idx)
                if cell.value is not None:
                    col_values.append(cell.value)

            if not col_values:
                continue

            if all(isinstance(v, (int, float)) for v in col_values):
                numeric_cols.append(col_idx)
            elif all(isinstance(v, str) for v in col_values):
                text_cols.append(col_idx)

        total_data_cells = sum(
            1 for row in ws.iter_rows() for cell in row if cell.value is not None
        )

        return {
            "empty": False,
            "numeric_columns": len(numeric_cols),
            "text_columns": len(text_cols),
            "date_columns": len(date_cols),
            "total_data_cells": total_data_cells,
        }

    def _extract_number_formats(self, ws) -> Dict[str, Any]:
        """Extract number format information."""
        formats_used: Dict[str, int] = {}
        max_rows = min(ws.max_row, 50)
        max_cols = min(ws.max_column, 20)

        for row in ws.iter_rows(min_row=1, max_row=max_rows, max_col=max_cols):
            for cell in row:
                if cell.value is not None and cell.number_format:
                    fmt = cell.number_format
                    if fmt != "General":
                        formats_used[fmt] = formats_used.get(fmt, 0) + 1

        keys = list(formats_used.keys())

        return {
            "formats_used": keys[:10],
            "has_currency": any(
                any(sym in fmt for sym in ("$", "£", "€")) for fmt in keys
            ),
            "has_percentage": any("%" in fmt for fmt in keys),
            "has_date": any(
                any(ch in fmt.lower() for ch in ("d", "m", "y")) for fmt in keys
            ),
        }

    def _analyze_table_structure(self, ws) -> Dict[str, Any]:
        """Analyze table-like structure of the sheet."""
        return {
            "has_multiple_sheets": len(self.workbook.sheetnames) > 1,
            "sheet_count": len(self.workbook.sheetnames),
            "appears_tabular": ws.max_row > 1 and ws.max_column > 1,
            "density": self._calculate_density(ws),
        }

    def _calculate_density(self, ws) -> float:
        """Calculate data density (% of cells with data)."""
        total_cells = ws.max_row * ws.max_column
        if total_cells == 0:
            return 0.0

        filled_cells = sum(
            1 for row in ws.iter_rows() for cell in row if cell.value is not None
        )
        return round(filled_cells / total_cells * 100, 2)
