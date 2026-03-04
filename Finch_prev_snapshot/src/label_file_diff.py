

import argparse
import json
import re
from pathlib import Path
from typing import Dict, List, Tuple

import openpyxl
from openpyxl.chartsheet import Chartsheet

# ： Excel 
_READ_WORKBOOK_CACHE: Dict[Path, dict] = {}

# ---------- Helpers ----------

_addr_re = re.compile(r"^([A-Za-z]+)(\d+)$")


def addr_key(addr: str):
    """Sort key for addresses like A1, B10, AA3."""
    m = _addr_re.match(addr)
    if not m:
        return (10**9, 10**9)
    col_letters, row_str = m.groups()
    row = int(row_str)

    col_idx = 0
    for ch in col_letters.upper():
        col_idx = col_idx * 26 + (ord(ch) - ord("A") + 1)
    return (row, col_idx)


def read_sheet_to_maps(
    ws_formula,
    ws_values=None,
) -> Tuple[Dict[str, str], Dict[str, str], List[dict]]:
    addr2text: Dict[str, str] = {}
    addr2form: Dict[str, str] = {}

    for row in ws_formula.iter_rows(values_only=False):
        for cell in row:
            addr = cell.coordinate

            
            f_val = cell.value

            # /： ws_values 
            if ws_values is not None:
                v_cell = ws_values[addr]
                v_val = v_cell.value
            else:
                v_val = cell.value

            if f_val is None and v_val is None:
                continue

            text = ""
            formula = ""

            # addr2form 
            if isinstance(f_val, str) and f_val.startswith("="):
                formula = f_val

            # addr2text（，） 
            if v_val is not None:
                s = str(v_val).strip()
                if s:
                    text = s

            if text:
                addr2text[addr] = text
            if formula:
                addr2form[addr] = formula

    items: List[dict] = []
    for addr in sorted(set(addr2text.keys()) | set(addr2form.keys()), key=addr_key):
        # attention 
        item = {"address": addr, "text": addr2text.get(addr, "")}
        if addr in addr2form:
            item["formula"] = addr2form[addr]
        items.append(item)

    return addr2text, addr2form, items


def read_workbook(path: Path):
    # ， key 
    path = Path(path).resolve()

    
    cached = _READ_WORKBOOK_CACHE.get(path)
    if cached is not None:
        return cached

    
    wb_formula = openpyxl.load_workbook(path, data_only=False)
    
    wb_values = openpyxl.load_workbook(path, data_only=True)

    # ChartSheet，： cell diff 
    for ws in getattr(wb_formula, "_sheets", []):
        if isinstance(ws, Chartsheet):
            print(
                f"   !! read_workbook: skip chart sheet '{ws.title}' in {path.name} "
                f"(chart sheet has no cell grid; handled by chart diff only)."
            )

    sheet_maps = {}
    # wb_formula.worksheets Worksheet， Chartsheet， 
    for ws_formula in wb_formula.worksheets:
        title = ws_formula.title
        ws_values = wb_values[title] if title in wb_values.sheetnames else None

        addr2text, addr2form, items = read_sheet_to_maps(ws_formula, ws_values)
        sheet_maps[title] = (addr2text, addr2form, items)

    
    _READ_WORKBOOK_CACHE[path] = sheet_maps
    return sheet_maps




# ---------- Diff logic ----------

def diff_two_sheets(
    before_maps, after_maps
) -> Tuple[List[dict], List[dict], List[dict]]:
    """
    Given (addr2text, addr2form, items) for before/after, compute diff.
    Returns: before_items, after_items, diff_items
    """
    b_text, b_form, before_items = before_maps
    a_text, a_form, after_items = after_maps

    all_addrs = set(b_text.keys()) | set(a_text.keys()) | set(b_form.keys()) | set(a_form.keys())
    diff_items: List[dict] = []

    for addr in sorted(all_addrs, key=addr_key):
        tb = b_text.get(addr, "")
        ta = a_text.get(addr, "")
        fb = b_form.get(addr, "")
        fa = a_form.get(addr, "")

        # -------- text （）--------- 
        same_text = (tb == ta)

        if not same_text:
            # text ， 6 
            try:
                tb_float = float(tb)
                ta_float = float(ta)
            except ValueError:
                # ， same_text = False 
                pass
            else:
                if round(tb_float, 6) == round(ta_float, 6):
                    
                    same_text = True

        # -------- formula （）--------- 
        same_formula = (fb == fa)

        # text formula => diff 
        if same_text and same_formula:
            continue

        # => diff 
        diff_items.append(
            {
                "address": addr,
                "text_before": tb,
                "text_after": ta,
                "formula_before": fb,
                "formula_after": fa,
            }
        )

    return before_items, after_items, diff_items




def build_workbook_diff(before_path: Path, after_path: Path) -> dict:
    """
    Build a sheet-level diff between two workbooks.
    """
    before_wb = read_workbook(before_path)
    after_wb = read_workbook(after_path)

    all_sheets = set(before_wb.keys()) | set(after_wb.keys())
    out = {
        "before_file": str(before_path),
        "after_file": str(after_path),
        "before_sheets": {},   # ： sheet before 
        "after_sheets": {},    # ： sheet after 
        "sheets_diff": {},
    }

    for sheet in sorted(all_sheets):
        if sheet in before_wb:
            before_maps = before_wb[sheet]
        else:
            before_maps = ({}, {}, [])

        if sheet in after_wb:
            after_maps = after_wb[sheet]
        else:
            after_maps = ({}, {}, [])

        before_items, after_items, diff_items = diff_two_sheets(before_maps, after_maps)

        # diff， before/after 
        out["before_sheets"][sheet] = before_items
        out["after_sheets"][sheet] = after_items

        # sheet sheets_diff 
        if not diff_items and sheet in before_wb and sheet in after_wb:
            # completely identical sheet, skip in sheets_diff
            continue

        out["sheets_diff"][sheet] = {
            "before": before_items,
            "after": after_items,
            "diff": diff_items,
        }

    return out


# ---------- CLI ----------

def main():
    parser = argparse.ArgumentParser(
        description="Compare two Excel files sheet-by-sheet and save diffs as JSON."
    )
    parser.add_argument("before", type=Path, help="Path to the 'before' Excel file")
    parser.add_argument("after", type=Path, help="Path to the 'after' Excel file")
    parser.add_argument(
        "-o",
        "--output",
        type=Path,
        default=None,
        help="Optional output JSON file path. "
             "If omitted, saves to ./label_diff/<after_file_name>_diff.json",
    )
    args = parser.parse_args()

    before = args.before.resolve()
    after = args.after.resolve()

    diff = build_workbook_diff(before, after)

    if args.output is None:
        out_dir = Path.cwd() / "label_diff"
        out_dir.mkdir(parents=True, exist_ok=True)
        out_path = out_dir / f"{after.stem}_diff.json"
    else:
        out_path = args.output
        out_path.parent.mkdir(parents=True, exist_ok=True)

    with out_path.open("w", encoding="utf-8") as f:
        json.dump(diff, f, ensure_ascii=False, indent=2)

    print(f"[DONE] Diff saved to {out_path}")


if __name__ == "__main__":
    main()
