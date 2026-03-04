# gpt_judge_eval_with_sd.py
"""
CLI entrypoint for running GPT-based judgment over a root directory of cases.

Usage:
    python gpt_judge_eval.py \
        --root-dir "/path/to/cases" \
        --out-excel "gpt_judge_results.xlsx"
"""

import argparse
import json
import logging
import time
from pathlib import Path
from typing import Dict, Any, List, Optional

import pandas as pd

from src.config import PROMPT_MAX_CHARS
from .gpt_judge_context import (
    build_case_judge_context,
    truncate_prompt_to_max,
    read_task_classification,
)
from .gpt_judge_helpers import call_gpt_judge, ensure_output_xlsx
from .gpt_judge_context import read_optional_text_file  # if needed
from openpyxl import load_workbook

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s %(levelname)s %(message)s",
    filename="logs/gpt_judge_eval.log",
    filemode="a",
)


def detect_human_label(case_dir: Path) -> str:
    """Return 'pass', 'fail', or 'unknown' based on pass.txt / fail.txt."""
    if (case_dir / "pass.txt").is_file():
        return "pass"
    if (case_dir / "fail.txt").is_file():
        return "fail"
    return "unknown"


def compute_judge_status(llm_label: str, human_label: str) -> str:
    """Return 'agree', 'disagree', or 'unknown' based on LLM vs human labels."""
    if llm_label not in ("pass", "fail"):
        return "unknown"
    if human_label not in ("pass", "fail"):
        return "unknown"
    return "agree" if llm_label == human_label else "disagree"


def main():
    parser = argparse.ArgumentParser(
        description="Use gpt-5-mini as judge to compare model output.xlsx with answer.xlsx."
    )
    parser.add_argument(
        "--root-dir",
        type=Path,
        required=True,
        help="Root directory, for example: C:\\Users\\hadong\\Downloads\\LLM eval test data-20251110T091736Z-1-001\\LLM eval test data",
    )
    parser.add_argument(
        "--out-excel",
        type=Path,
        default=Path("gpt_judge_results.xlsx"),
        help="Output Excel result path (default: ./gpt_judge_results.xlsx)",
    )
    args = parser.parse_args()

    root_dir = args.root_dir.resolve()
    out_excel = args.out_excel.resolve()

    if not root_dir.is_dir():
        raise NotADirectoryError(f"Root dir not found: {root_dir}")

    results: List[Dict[str, Any]] = []
    out_excel.parent.mkdir(parents=True, exist_ok=True)

    ctx_dir = Path.cwd() / "llm_judge_ctx"
    ctx_dir.mkdir(parents=True, exist_ok=True)

    for case_dir in sorted(root_dir.iterdir()):
        if not case_dir.is_dir():
            continue

        folder_name = case_dir.name
        print(f"[CASE] {folder_name}")
        logging.info("Processing case: %s", folder_name)

        
        input_path = case_dir / "input.xlsx"
        answer_path = case_dir / "answer.xlsx"

        # ========= ：ensure_output_xlsx try/except， ========= 
        try:
            output_path = ensure_output_xlsx(case_dir)
        except Exception as e:
            msg = f"ensure_output_xlsx error in case '{folder_name}': {e}"
            print(f"[WARN] {msg}")
            logging.warning(msg)

            human_label = detect_human_label(case_dir)
            results.append(
                {
                    "folder": folder_name,
                    "llm_label": "",
                    "llm_score": None,
                    "llm_reason": "",
                    "human_label": human_label,
                    "judge_status": "unknown",
                    "llm_call_status": "error",
                    "error_message": msg,
                    "prompt_chars": None,
                    "task_classification": "modify",  # output modify ， modify 
                }
            )
            pd.DataFrame(results).to_excel(out_excel, index=False)
            # case， 
            continue

        query_path = case_dir / "query.txt"

        # ：modify / generate（ property.json task_classification） 
        file_mode = read_task_classification(case_dir)
        
                # NEW: task_classification generate， 
        # input answer / output sheet name ， modify 
        if file_mode == "generate":
            try:
                if input_path.is_file() and answer_path.is_file() and output_path.is_file():
                    wb_input = load_workbook(filename=str(input_path), data_only=False)
                    wb_answer = load_workbook(filename=str(answer_path), data_only=False)
                    wb_output = load_workbook(filename=str(output_path), data_only=False)

                    input_sheets = {ws.title for ws in wb_input.worksheets}
                    answer_sheets = {ws.title for ws in wb_answer.worksheets}
                    output_sheets = {ws.title for ws in wb_output.worksheets}

                    has_input_answer_overlap = bool(input_sheets & answer_sheets)
                    has_input_output_overlap = bool(input_sheets & output_sheets)

                    if has_input_answer_overlap and has_input_output_overlap:
                        print(
                            f"[MODE] Case '{folder_name}' is classified as generate but "
                            f"input/answer/output share sheet names; switch to modify mode."
                        )
                        file_mode = "modify"
            except Exception as e:
                # ，， file_mode 
                print(f"[MODE][WARN] Failed to inspect workbooks for case '{folder_name}': {e}")
                
                
        if file_mode in ("generate", "qa"):
            # / QA ： query.txt ， 
            # answer.xlsx / output.xlsx （/） 
            required_paths = [query_path]
        else:  # modify 
            required_paths = [input_path, answer_path, output_path, query_path]


        missing = [p.name for p in required_paths if not p.is_file()]

        if missing:
            msg = f"Missing files ({file_mode} task): {', '.join(missing)}"
            print(f"[WARN] {msg}")
            logging.warning("Case '%s' skipped due to missing files: %s", folder_name, ", ".join(missing))

            human_label = detect_human_label(case_dir)
            results.append(
                {
                    "folder": folder_name,
                    "llm_label": "",
                    "llm_score": None,
                    "llm_reason": "",
                    "human_label": human_label,
                    "judge_status": "unknown",
                    "llm_call_status": "error",
                    "error_message": msg,
                    "prompt_chars": None,
                    "task_classification": file_mode,
                }
            )
            pd.DataFrame(results).to_excel(out_excel, index=False)
            continue

        # ========= ： query.txt try/except ========= 
        try:
            query_text = query_path.read_text(encoding="utf-8", errors="ignore").strip()
        except Exception as e:
            msg = f"Failed to read query.txt in case '{folder_name}': {e}"
            print(f"[WARN] {msg}")
            logging.warning(msg)

            human_label = detect_human_label(case_dir)
            results.append(
                {
                    "folder": folder_name,
                    "llm_label": "",
                    "llm_score": None,
                    "llm_reason": "",
                    "human_label": human_label,
                    "judge_status": "unknown",
                    "llm_call_status": "error",
                    "error_message": msg,
                    "prompt_chars": None,
                    "task_classification": file_mode,
                }
            )
            pd.DataFrame(results).to_excel(out_excel, index=False)
            continue

        # ========= ： try/except， ========= 
        try:
            ctx = build_case_judge_context(
                input_path,
                answer_path,
                output_path,
                query_text,
                file_mode=file_mode,
            )

            prompt = ctx["prompt"]
            prompt_chars = ctx["prompt_chars"]

        except Exception as e:
            msg = f"build_case_judge_context error in case '{folder_name}': {e}"
            print(f"[WARN] {msg}")
            logging.warning(msg)

            human_label = detect_human_label(case_dir)
            results.append(
                {
                    "folder": folder_name,
                    "llm_label": "",
                    "llm_score": None,
                    "llm_reason": "",
                    "human_label": human_label,
                    "judge_status": "unknown",
                    "llm_call_status": "error",
                    "error_message": msg,
                    "prompt_chars": None,
                    "task_classification": file_mode,
                }
            )
            pd.DataFrame(results).to_excel(out_excel, index=False)
            continue

        # ========= prompt ，， ========= 
        if prompt_chars is not None and prompt_chars > PROMPT_MAX_CHARS:
            msg = (
                f"Prompt too long in case '{folder_name}': "
                f"{prompt_chars} > MAX_PROMPT_CHARS={PROMPT_MAX_CHARS}, will be truncated."
            )
            print(f"[WARN] {msg}")
            logging.warning(msg)

            # prompt 
            prompt = truncate_prompt_to_max(prompt, PROMPT_MAX_CHARS)
            prompt_chars = len(prompt)

            # ctx ， 
            ctx["prompt"] = prompt
            ctx["prompt_chars"] = prompt_chars

        # ，（）ctx 
        ctx_path = ctx_dir / f"{folder_name}.json"
        with ctx_path.open("w", encoding="utf-8") as f:
            json.dump(ctx, f, ensure_ascii=False, indent=2)


        # bundle input/、answer/、output/ （ generate / qa ） 
        folder_image_paths: List[Path] = []
        try:
            bundle = ctx.get("bundle", {})
            folder_images_info = bundle.get("folder_images") or {}
            for role, paths in folder_images_info.items():
                if not isinstance(paths, list):
                    continue
                for s in paths:
                    try:
                        p = Path(s)
                        folder_image_paths.append(p)
                    except Exception:
                        continue
        except Exception:
            folder_image_paths = []

        
        max_retries = 3
        attempt = 0
        judge_res: Dict[str, Any] = {
            "score": None,
            "detailed_analysis": "",
            "llm_call_status": "error",
            "error_message": "LLM not called yet",
        }

        
        image_exts = {".png", ".jpg", ".jpeg", ".bmp"}

        input_img_path: Optional[Path] = None
        answer_img_path: Optional[Path] = None
        output_img_path: Optional[Path] = None
        extra_output_images: List[Path] = []

        if file_mode in ("generate", "qa"):
            # / QA ： input/answer/output ， 
            # extra_output_images ，。 

            answer_main = case_dir / "_answer.png"
            if answer_main.is_file():
                answer_img_path = answer_main

            output_main = case_dir / "_output.png"
            if output_main.is_file():
                output_img_path = output_main

            answer_extras: List[Path] = []
            output_extras: List[Path] = []
            other_images: List[Path] = []

            for p in sorted(case_dir.iterdir()):
                if not p.is_file():
                    continue
                if p.suffix.lower() not in image_exts:
                    continue

                lower_name = p.name.lower()

                if lower_name in ("_answer.png", "_output.png"):
                    continue

                if lower_name.startswith("answer"):
                    answer_extras.append(p)
                elif lower_name.startswith("output"):
                    output_extras.append(p)
                else:
                    other_images.append(p)

            ordered_images: List[Path] = []
            if answer_img_path:
                ordered_images.append(answer_img_path)
            ordered_images.extend(answer_extras)
            if output_img_path:
                ordered_images.append(output_img_path)
            ordered_images.extend(output_extras)
            ordered_images.extend(other_images)

            # input/、answer/、output/ 
            if folder_image_paths:
                ordered_images.extend(folder_image_paths)

            extra_output_images = ordered_images

            has_any_image = bool(extra_output_images)
            if has_any_image:
                print(
                    f"[CASE] {folder_name} -> visual judging ({file_mode}) with {len(extra_output_images)} image(s): "
                    f"{', '.join(p.name for p in extra_output_images)}"
                )
            else:
                print(f"[CASE] {folder_name} -> no images found, judging by text only.")

            # generate / qa ， input/answer/output 
            input_img_path = None
            answer_img_path = None
            output_img_path = None

        else:
            
            input_img_path = case_dir / "_input.png"
            answer_img_path = case_dir / "_answer.png"
            output_img_path = case_dir / "_output.png"

            if not input_img_path.is_file():
                input_img_path = None
            if not answer_img_path.is_file():
                answer_img_path = None
            if not output_img_path.is_file():
                output_img_path = None

            has_any_image = any([input_img_path, answer_img_path, output_img_path])
            if has_any_image:
                print(
                    f"[CASE] {folder_name} -> visual judging with images: "
                    f"input={bool(input_img_path)}, answer={bool(answer_img_path)}, output={bool(output_img_path)}"
                )
            else:
                print(f"[CASE] {folder_name} -> no images found, judging by text only.")

            extra_output_images = []
            for p in sorted(case_dir.iterdir()):
                if not p.is_file():
                    continue
                if p.suffix.lower() not in image_exts:
                    continue

                lower_name = p.name.lower()

                if lower_name in ("_input.png", "_answer.png", "_output.png"):
                    continue

                if lower_name.startswith(("_input_", "_answer_", "_output_")):
                    continue

                extra_output_images.append(p)

            if extra_output_images:
                print(
                    f"[CASE] {folder_name} -> found {len(extra_output_images)} extra image(s): "
                    f"{', '.join(p.name for p in extra_output_images)}"
                )

        while attempt < max_retries:
            attempt += 1
            print(f"[LLM] Calling judge for case {folder_name}, attempt {attempt}/{max_retries} ...")

            if file_mode in ("generate", "qa"):
                # & QA ： query + answer/output LLM（） 
                judge_res = call_gpt_judge(
                    prompt,
                    input_image=None,
                    answer_image=None,
                    output_image=None,
                    extra_output_images=extra_output_images,
                )
            else:
                # ： input/answer/output 
                judge_res = call_gpt_judge(
                    prompt,
                    input_image=input_img_path,
                    answer_image=answer_img_path,
                    output_image=output_img_path,
                    extra_output_images=extra_output_images,
                )


            llm_call_status = judge_res.get("llm_call_status", "error")
            if llm_call_status != "error":
                # error ， 
                break

            # ： 429 Too Many Requests， 70 
            err_msg = str(judge_res.get("error_message", ""))
            if "429" in err_msg or "Too Many Requests" in err_msg:
                if attempt < max_retries:
                    print(
                        f"[LLM] Got 429 Too Many Requests for case {folder_name}, "
                        f"waiting 70 seconds before retry ({attempt}/{max_retries})..."
                    )
                    time.sleep(70)
            # 429，：，（） 

        # error， 
        if judge_res.get("llm_call_status", "error") == "error":
            logging.warning(
                "LLM call failed for case '%s' after %d attempt(s): %s",
                folder_name,
                attempt,
                judge_res.get("error_message", ""),
            )

        # ========= =========

        score = judge_res["score"]
        llm_call_status = judge_res["llm_call_status"]

        # NEW: score， fail，score=0 
        if score is None:
            score = 0
            llm_label = "fail"
        else:
            llm_label = "pass" if score == 1 else "fail"


        human_label = detect_human_label(case_dir)
        judge_status = compute_judge_status(llm_label, human_label)

        results.append(
            {
                "folder": folder_name,
                "llm_label": llm_label,
                "llm_score": score,
                "llm_reason": judge_res["detailed_analysis"],
                "human_label": human_label,
                "judge_status": judge_status,
                "llm_call_status": llm_call_status,
                "error_message": judge_res["error_message"],
                "prompt_chars": prompt_chars,
                "llm_call_attempts": attempt,
                "task_classification": file_mode,
            }
        )

        pd.DataFrame(results).to_excel(out_excel, index=False)

    print(f"[DONE] Saved judge results to {out_excel}")
    logging.info("Saved judge results to %s", out_excel)




if __name__ == "__main__":
    main()
